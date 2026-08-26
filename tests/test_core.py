import csv
import json
import os
import tempfile
import unittest
from contextlib import closing
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

from tax_system.core import (
    EXPORT_DATA_COLUMNS,
    INVENTORY_COLUMNS,
    LEDGER_COLUMNS,
    LEDGER_IDENTITY_COLUMNS,
    LEDGER_POS_COLUMNS,
    TaxSystem,
    _to_date,
    suggest_ledger_items,
)


class LedgerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_csv(self, row):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader(); writer.writerow(row)
        return path

    def test_valid_ledger_roundtrip(self):
        row = dict(zip(LEDGER_COLUMNS, ["2026-01-01", "テスト", "てすと", "2000-01-01", "匿名住所", "000", "商品", "2", "100", "200", ""])); path = self.write_csv(row)
        import_id = self.app.import_ledger(path)
        self.assertEqual([], self.app.validate(import_id))
        output = self.root / "out.csv"; self.app.export(import_id, output)
        self.assertTrue(output.read_bytes().startswith(b"\xef\xbb\xbf"))
        self.assertIn(b"\r\n", output.read_bytes())

    def test_ledger_import_tolerates_missing_remarks_column(self):
        # some exports omit 備考 entirely (not just leave it blank) — the column itself
        # is absent from the header, not just empty.
        columns_without_remarks = LEDGER_COLUMNS[:-1]
        path = self.root / "ledger_no_remarks.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns_without_remarks, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow(dict(zip(columns_without_remarks, [
                "2026-01-01", "テスト", "てすと", "2000-01-01", "匿名住所", "000", "商品", "2", "100", "200",
            ])))
        import_id = self.app.import_ledger(path)
        records, total = self.app.get_records(import_id)
        self.assertEqual(1, total)
        self.assertEqual("", records[0]["data"]["備考"])
        self.assertEqual("商品", records[0]["data"]["商品名"])

    def test_error_blocks_formal_export(self):
        row = dict(zip(LEDGER_COLUMNS, ["2026-01-01", "", "", "", "", "", "商品", "2", "100", "999", ""])); path = self.write_csv(row)
        import_id = self.app.import_ledger(path)
        with self.assertRaises(ValueError): self.app.export(import_id, self.root / "blocked.csv")
        self.app.export(import_id, self.root / "preview.csv", preview=True)

    def test_get_records_sorts_by_numeric_column(self):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for amount in (300, 100, 200):
                writer.writerow(dict(zip(LEDGER_COLUMNS, [
                    "2026-01-01", f"名前{amount}", "", "", "住所", "000", "商品", "1", str(amount), str(amount), "",
                ])))
        import_id = self.app.import_ledger(path)

        asc, total = self.app.get_records(import_id, sort="金額", sort_dir="asc")
        self.assertEqual(["100", "200", "300"], [r["data"]["金額"] for r in asc])

        desc, _ = self.app.get_records(import_id, sort="金額", sort_dir="desc")
        self.assertEqual(["300", "200", "100"], [r["data"]["金額"] for r in desc])

    def test_get_records_can_filter_to_a_single_row(self):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for amount in (300, 100, 200):
                writer.writerow(dict(zip(LEDGER_COLUMNS, [
                    "2026-01-01", f"名前{amount}", "", "", "住所", "000", "商品", "1", str(amount), str(amount), "",
                ])))
        import_id = self.app.import_ledger(path)
        all_rows, _ = self.app.get_records(import_id)
        target_row_no = all_rows[1]["row_no"]

        rows, total = self.app.get_records(import_id, row_no=target_row_no)

        self.assertEqual(1, total)
        self.assertEqual(1, len(rows))
        self.assertEqual(target_row_no, rows[0]["row_no"])

    def test_get_records_can_filter_to_a_list_of_rows(self):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for amount in (300, 100, 200):
                writer.writerow(dict(zip(LEDGER_COLUMNS, [
                    "2026-01-01", f"名前{amount}", "", "", "住所", "000", "商品", "1", str(amount), str(amount), "",
                ])))
        import_id = self.app.import_ledger(path)
        all_rows, _ = self.app.get_records(import_id)
        targets = [all_rows[0]["row_no"], all_rows[2]["row_no"]]

        rows, total = self.app.get_records(import_id, row_no=targets)

        self.assertEqual(2, total)
        self.assertEqual(set(targets), {r["row_no"] for r in rows})

    def test_export_can_filter_to_a_single_month(self):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for date_str, name in [("2026-04-05", "A"), ("2026-04-20", "B"), ("2026-05-03", "C")]:
                writer.writerow(dict(zip(LEDGER_COLUMNS, [
                    date_str, name, "", "1990-01-01", "住所", "000", "商品", "1", "1000", "1000", "",
                ])))
        import_id = self.app.import_ledger(path)

        output = self.root / "april.csv"
        self.app.export(import_id, output, month="2026-04")
        with output.open(encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
        self.assertEqual({"A", "B"}, {r["名前"] for r in rows})

        with self.assertRaises(ValueError):
            self.app.export(import_id, self.root / "none.csv", month="2026-06")

    def test_header_echoed_as_data_row_is_skipped(self):
        # concatenating multiple exports into one CSV can leave a duplicate header
        # line in the middle, which must not be imported as a real record.
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow(dict(zip(LEDGER_COLUMNS, ["2026-01-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""])))
            writer.writerow(dict(zip(LEDGER_COLUMNS, LEDGER_COLUMNS)))  # duplicated header, no real data
            writer.writerow(dict(zip(LEDGER_COLUMNS, ["2026-01-02", "B", "びー", "1991-01-01", "住所B", "000", "商品B", "1", "2000", "2000", ""])))
        import_id = self.app.import_ledger(path)
        records, total = self.app.get_records(import_id)
        self.assertEqual(2, total)
        self.assertNotIn("日時", [r["data"].get("日時") for r in records])

    def test_delete_import_removes_records(self):
        row = dict(zip(LEDGER_COLUMNS, ["2026-01-01", "テスト", "てすと", "2000-01-01", "匿名住所", "000", "商品", "2", "100", "200", ""]))
        path = self.write_csv(row)
        import_id = self.app.import_ledger(path)
        self.app.delete_import(import_id)
        self.assertIsNone(self.app.get_import(import_id))
        records, total = self.app.get_records(import_id)
        self.assertEqual(0, total)
        with self.assertRaises(ValueError):
            self.app.delete_import(import_id)

    def test_rename_import_source_updates_display_name(self):
        row = dict(zip(LEDGER_COLUMNS, ["2026-01-01", "テスト", "てすと", "2000-01-01", "匿名住所", "000", "商品", "2", "100", "200", ""]))
        import_id = self.app.import_ledger(self.write_csv(row))
        self.app.rename_import_source(import_id, "元のファイル名.csv")
        self.assertEqual("元のファイル名.csv", self.app.get_import(import_id)["source_name"])


class AllocationTests(unittest.TestCase):
    PURCHASE_HEADERS = ["年月日", "品目", "数量", "相手方名", "代価"]
    SALE_HEADERS = ["年月日", "数量"]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, row):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader(); writer.writerow(row)
        return path

    def write_comparison(self, purchase_row, sale_row, name="comparison.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "輸出販売"
        headers = self.PURCHASE_HEADERS + self.SALE_HEADERS
        for col, value in enumerate(headers, 1):
            ws.cell(2, col).value = value
        for col, value in enumerate(purchase_row + sale_row, 1):
            ws.cell(3, col).value = value
        path = self.root / name
        wb.save(path)
        return path

    def test_matching_purchase_is_found_automatically(self):
        ledger_row = dict(zip(LEDGER_COLUMNS, [
            "2026-05-01", "山田太郎", "やまだたろう", "1990-01-01", "匿名住所", "000",
            "テストカード", "2", "10000", "20000", "",
        ]))
        self.app.import_ledger(self.write_ledger(ledger_row))
        comparison_id = self.app.import_comparison(
            self.write_comparison(["2026-05-01", "テストカード", 2, "山田太郎", 20000], ["2026-06-01", 2])
        )
        self.assertEqual([], self.app.validate(comparison_id))

    def test_missing_purchase_blocks_export(self):
        ledger_row = dict(zip(LEDGER_COLUMNS, [
            "2026-05-01", "山田太郎", "やまだたろう", "1990-01-01", "匿名住所", "000",
            "テストカード", "2", "10000", "20000", "",
        ]))
        self.app.import_ledger(self.write_ledger(ledger_row))
        comparison_id = self.app.import_comparison(
            self.write_comparison(["2026-05-01", "別の商品", 2, "山田太郎", 20000], ["2026-06-01", 2])
        )
        checks = self.app.validate(comparison_id)
        self.assertTrue(any(c.code == "ALLOCATION_NOT_FOUND" for c in checks))
        with self.assertRaises(ValueError):
            self.app.export(comparison_id, self.root / "blocked.xlsx", template_id=None)

    def test_ledger_record_is_not_reused_for_two_sales(self):
        ledger_row = dict(zip(LEDGER_COLUMNS, [
            "2026-05-01", "山田太郎", "やまだたろう", "1990-01-01", "匿名住所", "000",
            "テストカード", "2", "10000", "20000", "",
        ]))
        self.app.import_ledger(self.write_ledger(ledger_row))
        first_id = self.app.import_comparison(
            self.write_comparison(["2026-05-01", "テストカード", 2, "山田太郎", 20000], ["2026-06-01", 2], name="first.xlsx")
        )
        second_id = self.app.import_comparison(
            self.write_comparison(["2026-05-01", "テストカード", 2, "山田太郎", 20000], ["2026-06-02", 2], name="second.xlsx")
        )
        self.assertEqual([], self.app.validate(first_id))
        checks = self.app.validate(second_id)
        self.assertTrue(any(c.code == "ALLOCATION_NOT_FOUND" for c in checks))


class LedgerCompletionTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_inventory(self):
        path = self.root / "inventory.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=INVENTORY_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow({"商品名": "テストカードA", "仕入れ原価": "5000", "在庫数": "10"})
            writer.writerow({"商品名": "テストカードB", "仕入れ原価": "3000", "在庫数": "5"})
        return path

    def import_inventory_for(self, month):
        self.app.import_inventory(self.write_inventory(), as_of=month)

    def write_ledger_lump(self, total):
        path = self.root / "ledger.csv"
        row = dict(zip(LEDGER_COLUMNS, [
            "2026-05-01", "鈴木一郎", "すずきいちろう", "1985-01-01", "匿名住所", "000",
            "", "", "", str(total), "",
        ]))
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader(); writer.writerow(row)
        return path

    def write_comparison(self, rows, name="comparison.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "輸出販売"
        headers = ["年月日", "品目", "数量", "相手方名", "代価", "年月日", "数量"]
        for col, value in enumerate(headers, 1):
            ws.cell(2, col).value = value
        for r, (purchase_row, sale_row) in enumerate(rows, 3):
            for col, value in enumerate(purchase_row + sale_row, 1):
                ws.cell(r, col).value = value
        path = self.root / name
        wb.save(path)
        return path

    def test_inventory_for_month_uses_exact_basis_month_not_latest(self):
        path1 = self.root / "inv_may.csv"
        with path1.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=INVENTORY_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow({"商品名": "テストカードA", "仕入れ原価": "9999", "在庫数": "1"})
        self.app.import_inventory(path1, as_of="2026-05")

        path2 = self.root / "inv_april.csv"
        with path2.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=INVENTORY_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow({"商品名": "テストカードA", "仕入れ原価": "1111", "在庫数": "1"})
        self.app.import_inventory(path2, as_of="2026-04")

        # each month must use its OWN basis-month inventory, never the newer one
        self.assertEqual(1111, self.app._inventory_for_month("2026-04")["テストカードA"])
        self.assertEqual(9999, self.app._inventory_for_month("2026-05")["テストカードA"])
        self.assertEqual({}, self.app._inventory_for_month("2026-06"))
        self.assertEqual({}, self.app._inventory_for_month(None))

    def test_breakdown_uses_comparison_and_manual_fill(self):
        self.import_inventory_for("2026-05")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(11000))
        self.app.import_comparison(self.write_comparison([
            (["2026-05-01", "テストカードA", 1, "鈴木一郎", 5000], ["2026-06-01", 1]),
            (["2026-05-01", "テストカードB", 1, "鈴木一郎", 3000], ["2026-06-02", 1]),
        ]))

        breakdown = self.app.propose_ledger_breakdown(ledger_id)
        self.assertEqual(1, len(breakdown))
        entry = breakdown[0]
        self.assertEqual("2026-05", entry["month"])
        self.assertEqual(2, len(entry["known_items"]))
        self.assertAlmostEqual(8000, sum(i["amount"] for i in entry["known_items"]))
        self.assertFalse(entry["resolved"])
        self.assertAlmostEqual(3000, entry["remainder"])

        self.app.add_ledger_item(ledger_id, entry["row_no"], "テストカードB", 1)
        resolved_breakdown = self.app.propose_ledger_breakdown(ledger_id)
        self.assertTrue(resolved_breakdown[0]["resolved"])

        output = self.root / "completed.csv"
        self.app.export_completed_ledger(ledger_id, output)
        with output.open(encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
        self.assertEqual(3, len(rows))
        self.assertAlmostEqual(11000, sum(float(r["金額"]) for r in rows))

    def test_export_blocked_when_unresolved(self):
        self.import_inventory_for("2026-05")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(11000))
        self.app.import_comparison(self.write_comparison([
            (["2026-05-01", "テストカードA", 1, "鈴木一郎", 5000], ["2026-06-01", 1]),
        ]))
        with self.assertRaises(ValueError):
            self.app.export_completed_ledger(ledger_id, self.root / "blocked.csv")

    def test_breakdown_filters_by_target_month(self):
        self.import_inventory_for("2026-05")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(11000))
        self.app.import_comparison(self.write_comparison([
            (["2026-05-01", "テストカードA", 1, "鈴木一郎", 5000], ["2026-06-01", 1]),
        ]))

        self.assertEqual(1, len(self.app.propose_ledger_breakdown(ledger_id, month="2026-05")))
        self.assertEqual(0, len(self.app.propose_ledger_breakdown(ledger_id, month="2026-04")))

    def test_breakdown_prefers_feature_over_generic_product_name(self):
        # 品目 is a generic category (e.g. "ワンピースカード"); 特徴 is the specific card.
        # The breakdown must report the specific card, matching fill_comparison_purchase_from_ledger.
        self.import_inventory_for("2026-05")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(5000))
        wb = Workbook()
        ws = wb.active
        ws.title = "輸出販売"
        headers = ["年月日", "品目", "特徴", "数量", "相手方名", "代価", "年月日", "数量"]
        for col, value in enumerate(headers, 1):
            ws.cell(2, col).value = value
        ws.cell(3, 1).value = "2026-05-01"
        ws.cell(3, 2).value = "ワンピースカード"
        ws.cell(3, 3).value = "テストカードA"
        ws.cell(3, 4).value = 1
        ws.cell(3, 5).value = "鈴木一郎"
        ws.cell(3, 6).value = 5000
        ws.cell(3, 7).value = "2026-06-01"
        ws.cell(3, 8).value = 1
        path = self.root / "comparison.xlsx"
        wb.save(path)
        self.app.import_comparison(path)

        breakdown = self.app.propose_ledger_breakdown(ledger_id)

        self.assertEqual(1, len(breakdown[0]["known_items"]))
        self.assertEqual("テストカードA", breakdown[0]["known_items"][0]["product"])

    def test_breakdown_falls_back_to_manually_linked_amount_without_inventory_cost(self):
        # when the specific card has no 期末在庫表 cost on file, the breakdown should still
        # count the amount the user manually prorated when linking, instead of ignoring it.
        ledger_id = self.app.import_ledger(self.write_ledger_lump(50000))
        comparison_id = self.app.import_comparison(self.write_comparison([
            (["", "モンキー・D・ルフィ", "", "", ""], ["2026-06-01", 1]),
        ]))
        records, _ = self.app.get_records(ledger_id)
        ledger_record_id = records[0]["id"]
        records, _ = self.app.get_records(comparison_id)
        row_no = records[0]["row_no"]

        self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", row_no, ledger_record_id,
                                                    qty=1, amount=25000)

        breakdown = self.app.propose_ledger_breakdown(ledger_id)

        self.assertEqual(1, len(breakdown[0]["known_items"]))
        item = breakdown[0]["known_items"][0]
        self.assertEqual("モンキー・D・ルフィ", item["product"])
        self.assertIsNone(item["unit_cost"])
        self.assertEqual(25000, item["amount"])
        self.assertEqual(25000, breakdown[0]["remainder"])

    def test_add_ledger_item_rejects_non_matching_month_inventory(self):
        # inventory is only available for April; the ledger row is from May, so no
        # inventory should be usable for it, even though a (different-month) snapshot exists.
        self.import_inventory_for("2026-04")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(11000))
        breakdown = self.app.propose_ledger_breakdown(ledger_id)
        row_no = breakdown[0]["row_no"]
        self.assertFalse(breakdown[0]["inventory_available"])
        with self.assertRaises(ValueError):
            self.app.add_ledger_item(ledger_id, row_no, "テストカードA", 1)

    def test_suggest_ledger_completion_never_invents_products(self):
        self.import_inventory_for("2026-05")
        ledger_id = self.app.import_ledger(self.write_ledger_lump(11000))
        with patch("openai.OpenAI") as mock_openai_cls:
            mock_client = MagicMock()
            mock_openai_cls.return_value = mock_client
            mock_response = MagicMock()
            mock_response.choices[0].message.content = json.dumps({
                "items": [
                    {"product": "テストカードA", "qty": 1},
                    {"product": "存在しないカード", "qty": 1},
                ]
            })
            mock_client.chat.completions.create.return_value = mock_response
            with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
                row_no = self.app.propose_ledger_breakdown(ledger_id)[0]["row_no"]
                items = self.app.suggest_ledger_completion(ledger_id, row_no)
        self.assertEqual([{"product": "テストカードA", "qty": 1, "unit_cost": 5000, "amount": 5000}], items)


class BuildComparisonTests(unittest.TestCase):
    HEADERS = ["年月日", "品目", "数量", "相手方名", "代価", "年月日", "数量"]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, rows):
        path = self.root / "ledger.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_COLUMNS, row)))
        return path

    def write_export_data(self, rows):
        path = self.root / "export_data.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=EXPORT_DATA_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(EXPORT_DATA_COLUMNS, row)))
        return path

    def register_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "輸出販売"
        for col, value in enumerate(self.HEADERS, 1):
            ws.cell(2, col).value = value
        path = self.root / "template.xlsx"
        wb.save(path)
        return self.app.register_template("comparison", path, "test", "v1", "2026-01-01")

    def test_fifo_matches_oldest_purchase_first(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""],
            ["2026-02-01", "Bさん", "びー", "1991-01-01", "匿名住所", "000", "カードX", "1", "1200", "1200", ""],
        ]))
        export_id = self.app.import_export_data(self.write_export_data([
            ["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"],
            ["2026-03-02", "カードX", "2200", "1", "2200", "海外顧客2", "銀行振込", "JPY"],
        ]))
        template_id = self.register_template()

        result = self.app.build_comparison(export_id, template_id)
        self.assertEqual(2, result["total"])
        self.assertEqual(0, result["unmatched"])

        self.assertEqual([], self.app.validate(result["import_id"]))

        allocations = self.app.allocate(result["import_id"])
        by_row = {a["row_no"]: a for a in allocations}
        self.assertEqual("Aさん", by_row[2]["ledger"]["name"])
        self.assertEqual("Bさん", by_row[3]["ledger"]["name"])

    def test_unmatched_row_flagged_for_manual_allocation(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードY", "1", "1000", "1000", ""],
        ]))
        export_id = self.app.import_export_data(self.write_export_data([
            ["2026-03-01", "カードZ", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"],
        ]))
        template_id = self.register_template()

        result = self.app.build_comparison(export_id, template_id)
        self.assertEqual(1, result["unmatched"])

        checks = self.app.validate(result["import_id"])
        self.assertTrue(any(c.code == "ALLOCATION_NOT_FOUND" for c in checks))

    def test_manual_export_entry_feeds_build_comparison(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""],
        ]))
        export_id = self.app.record_export_entry([{
            "年月日": "2026-03-01", "品名": "カードX", "金額": 2000, "数量": 1, "小計": 2000,
            "相手方名": "海外顧客1", "支払方法": "銀行振込", "通貨": "JPY", "英語名": "Card X",
        }], "手入力（テスト）")
        template_id = self.register_template()

        result = self.app.build_comparison(export_id, template_id)
        self.assertEqual(1, result["total"])
        self.assertEqual(0, result["unmatched"])
        self.assertEqual([], self.app.validate(result["import_id"]))


class RawLedgerImportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_pos_csv(self, rows):
        path = self.root / "pos.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_POS_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_POS_COLUMNS, row)))
        return path

    def write_identity_xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        for col, value in enumerate(LEDGER_IDENTITY_COLUMNS, 1):
            ws.cell(1, col).value = value
        for r, row in enumerate(rows, 2):
            for col, value in enumerate(row, 1):
                ws.cell(r, col).value = value
        path = self.root / "identity.xlsx"
        wb.save(path)
        return path

    def test_pos_import_maps_fields_and_skips_unapproved(self):
        result = self.app.import_ledger_pos(self.write_pos_csv([
            ["1", "承認済み", "2026-04-01 10:00:00", "1", "テスト太郎", "明細なし", "カードA", "1", "1000", "1000", "備考1", "全体1"],
            ["2", "却下", "2026-04-02 10:00:00", "1", "テスト次郎", "明細なし", "カードB", "1", "500", "500", "", ""],
        ]))
        self.assertEqual(1, result["imported"])
        self.assertEqual(1, result["skipped"])
        records, _ = self.app.get_records(result["import_id"])
        data = records[0]["data"]
        self.assertEqual("テスト太郎", data["名前"])
        self.assertEqual("カードA", data["商品名"])
        self.assertEqual("", data["ふりがな"])
        self.assertIn("備考1", data["備考"])
        self.assertIn("全体1", data["備考"])

    def test_pos_import_tolerates_trailing_comma_in_header(self):
        # some POS exports add a trailing comma to every line (an extra blank column),
        # which must not cause a strict header mismatch.
        path = self.root / "pos_trailing_comma.csv"
        header = ",".join(LEDGER_POS_COLUMNS) + ",\r\n"
        row = "1,承認済み,2026-04-01 10:00:00,1,テスト太郎,明細なし,カードA,1,1000,1000,,,\r\n"
        path.write_text("﻿" + header + row, encoding="utf-8")
        result = self.app.import_ledger_pos(path)
        self.assertEqual(1, result["imported"])
        records, _ = self.app.get_records(result["import_id"])
        self.assertEqual("カードA", records[0]["data"]["商品名"])

    def test_pos_import_defers_multi_item_transactions_to_ledger_completion(self):
        # a second (and later) card in the same purchase is recorded as a continuation
        # row: 履歴ID/状態/日時/ユーザーID/氏名 blank, only カード名/数量/単価 filled.
        # The single 商品名/個数/単価 slot can't represent both items, so both are left
        # blank for 内訳復元 to reconstruct from 相対表, instead of silently dropping
        # the second item while keeping the first as if it were the whole purchase.
        result = self.app.import_ledger_pos(self.write_pos_csv([
            ["1", "承認済み", "2026-04-01 10:00:00", "1", "テスト太郎", "明細なし", "カードA", "1", "1000", "3530", "", ""],
            ["", "", "", "", "", "", "カードB", "1", "530", "", "", ""],
            ["2", "承認済み", "2026-04-02 10:00:00", "1", "テスト次郎", "明細なし", "カードC", "1", "2000", "2000", "", ""],
        ]))
        self.assertEqual(2, result["imported"])
        self.assertEqual(0, result["skipped"])
        records, _ = self.app.get_records(result["import_id"])
        by_name = {r["data"]["名前"]: r["data"] for r in records}
        self.assertEqual("", by_name["テスト太郎"]["商品名"])
        self.assertEqual("", by_name["テスト太郎"]["個数"])
        self.assertEqual("3530", by_name["テスト太郎"]["金額"])
        # single-item purchases are unaffected
        self.assertEqual("カードC", by_name["テスト次郎"]["商品名"])
        self.assertEqual("2000", by_name["テスト次郎"]["金額"])

    def test_identity_import_maps_fields_and_ignores_trailing_blank_column(self):
        result = self.app.import_ledger_identity(self.write_identity_xlsx([
            ["2026-04-01", "テスト太郎", "てすとたろう", "1990-01-01", "匿名住所", "000-0000", 5000],
        ]))
        self.assertEqual(1, result["imported"])
        records, _ = self.app.get_records(result["import_id"])
        data = records[0]["data"]
        self.assertEqual("てすとたろう", data["ふりがな"])
        self.assertEqual("", data["商品名"])
        self.assertEqual(5000, data["金額"])

    def write_identity_csv(self, rows, with_email):
        columns = LEDGER_IDENTITY_COLUMNS[:-1] + (["メールアドレス"] if with_email else []) + LEDGER_IDENTITY_COLUMNS[-1:]
        path = self.root / "identity.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh, lineterminator="\r\n")
            writer.writerow(columns)
            for row in rows:
                writer.writerow(row)
        return path

    def test_identity_csv_import_without_email_column(self):
        result = self.app.import_ledger_identity(self.write_identity_csv(
            [["2026-04-01", "テスト太郎", "てすとたろう", "1990-01-01", "匿名住所", "000-0000", "5000"]],
            with_email=False,
        ))
        self.assertEqual(1, result["imported"])
        records, _ = self.app.get_records(result["import_id"])
        self.assertEqual("テスト太郎", records[0]["data"]["名前"])

    def test_identity_csv_import_ignores_email_column(self):
        result = self.app.import_ledger_identity(self.write_identity_csv(
            [["2026-04-01", "テスト太郎", "てすとたろう", "1990-01-01", "匿名住所", "000-0000", "taro@example.com", "5000"]],
            with_email=True,
        ))
        self.assertEqual(1, result["imported"])
        records, _ = self.app.get_records(result["import_id"])
        data = records[0]["data"]
        self.assertEqual("テスト太郎", data["名前"])
        self.assertEqual("000-0000", data["電話番号"])
        self.assertEqual("5000", data["金額"])

    def test_import_auto_recognizes_identity_csv_with_email_column(self):
        path = self.write_identity_csv(
            [["2026-04-01", "テスト太郎", "てすとたろう", "1990-01-01", "匿名住所", "000-0000", "taro@example.com", "5000"]],
            with_email=True,
        )
        result = self.app.import_auto(path)
        self.assertEqual("ledger", result["kind"])
        self.assertEqual("古物台帳（本人確認データ）", result["label"])


class AutoImportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_csv(self, columns, rows, name):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(columns, row)))
        return path

    def test_detects_ledger_csv(self):
        path = self.write_csv(LEDGER_COLUMNS, [
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "unknown.csv")
        result = self.app.import_auto(path)
        self.assertEqual("ledger", result["kind"])

    def test_detects_inventory_csv(self):
        path = self.write_csv(INVENTORY_COLUMNS, [["商品X", "1000", "5"]], "unknown2.csv")
        result = self.app.import_auto(path)
        self.assertEqual("inventory", result["kind"])

    def test_detects_export_data_csv(self):
        path = self.write_csv(EXPORT_DATA_COLUMNS, [
            ["2026-05-01", "品X", "1000", "1", "1000", "海外顧客", "銀行振込", "JPY"],
        ], "unknown3.csv")
        result = self.app.import_auto(path)
        self.assertEqual("export_data", result["kind"])

    def test_detects_pos_csv(self):
        path = self.write_csv(LEDGER_POS_COLUMNS, [
            ["1", "承認済み", "2026-04-01 10:00:00", "1", "テスト太郎", "明細なし", "カードA", "1", "1000", "1000", "", ""],
        ], "unknown4.csv")
        result = self.app.import_auto(path)
        self.assertEqual("ledger", result["kind"])

    def test_ec_product_csv_extracts_relevant_columns(self):
        columns = ["商品番号", "商品名", "型番/品番", "JANコード", "カテゴリ", "サブカテゴリ",
                   "グループ", "販売価格", "在庫数", "仕入", "重量"]
        path = self.write_csv(columns, [
            ["155", "テストカード", "", "", "CAT", "SUB", "", "10800", "2", "8640", ""],
        ], "products_棚卸用.csv")
        result = self.app.import_auto(path)
        self.assertEqual("inventory", result["kind"])
        records, total = self.app.get_records(result["import_id"])
        self.assertEqual(1, total)
        self.assertEqual({
            "商品名": "テストカード", "仕入れ原価": "8640", "在庫数": "2",
            "カテゴリ": "CAT", "サブカテゴリ": "SUB", "グループ": "", "販売価格": "10800",
        }, records[0]["data"])

    def test_unrecognized_csv_raises(self):
        path = self.write_csv(["列A", "列B"], [["1", "2"]], "unknown5.csv")
        with self.assertRaises(ValueError):
            self.app.import_auto(path)


class MergeLedgerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, rows, name):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_COLUMNS, row)))
        return path

    def test_merge_allowed_even_with_unresolved_errors(self):
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-02", "", "", "", "", "", "商品B", "1", "2000", "2000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])
        self.assertEqual(2, result["total"])
        checks = self.app.validate(result["import_id"])
        self.assertTrue(any(c.code == "REQUIRED" for c in checks))
        self.assertEqual(3, len(self.app.list_imports()))  # originals kept, plus the new merged one

    def test_merge_combines_rows_and_flags_duplicates(self):
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
            ["2026-04-03", "B", "びー", "1991-01-01", "住所B", "000", "商品C", "1", "3000", "3000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])
        self.assertEqual(3, result["total"])
        self.assertEqual(1, len(result["duplicates"]))

        # the intentional duplicate row is also caught by the normal ledger validation
        checks = self.app.validate(result["import_id"])
        self.assertTrue(any(c.code == "DUPLICATE" for c in checks))

        output = self.app.export(result["import_id"], self.root / "merged.csv", preview=True)
        with output.open(encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
        self.assertEqual(3, len(rows))

    def test_auto_merge_requires_no_selection_and_excludes_prior_merges(self):
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "B", "びー", "1991-01-01", "住所B", "000", "商品B", "1", "2000", "2000", ""],
        ], "b.csv"))
        first_merge = self.app.auto_merge_ledger_imports()
        self.assertEqual(2, first_merge["total"])

        third = self.app.import_ledger(self.write_ledger([
            ["2026-06-01", "C", "しー", "1992-01-01", "住所C", "000", "商品C", "1", "3000", "3000", ""],
        ], "c.csv"))
        second_merge = self.app.auto_merge_ledger_imports()
        # re-derives from the three raw imports, not from the first merge result
        self.assertEqual(3, second_merge["total"])
        self.assertNotEqual(first_merge["import_id"], second_merge["import_id"])

    def test_auto_merge_fails_with_fewer_than_two_raw_imports(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        with self.assertRaises(ValueError):
            self.app.auto_merge_ledger_imports()

    def test_find_ledger_duplicates_and_delete_one(self):
        # same transaction recorded in two different raw sources (e.g. POS export and
        # identity-verification data): 日時・名前・商品名・個数・金額 match, but one row
        # carries extra identity fields the other lacks — this must NOT be silently lost.
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])
        self.assertEqual(1, len(result["duplicates"]))

        duplicates = self.app.find_ledger_duplicates(result["import_id"])
        self.assertEqual(1, len(duplicates))
        occurrences = duplicates[0]["occurrences"]
        self.assertEqual(2, len(occurrences))

        # the row with no identity details is recommended for deletion, the fuller one is not
        bare_row = next(o for o in occurrences if not o["data"]["ふりがな"])
        full_row = next(o for o in occurrences if o["data"]["ふりがな"])
        self.assertTrue(bare_row["recommended_delete"])
        self.assertFalse(full_row["recommended_delete"])

        self.app.delete_ledger_record(result["import_id"], bare_row["row_no"])

        self.assertEqual([], self.app.find_ledger_duplicates(result["import_id"]))
        _, total = self.app.get_records(result["import_id"])
        self.assertEqual(1, total)

    def test_find_ledger_duplicates_recommends_one_when_rows_are_byte_identical(self):
        # both rows tie on completeness but are otherwise fully identical — keeping
        # either one loses no information, so one is recommended for deletion. The user
        # can still override with dismiss_ledger_duplicate if it's a real coincidence.
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])

        occurrences = self.app.find_ledger_duplicates(result["import_id"])[0]["occurrences"]
        self.assertEqual(1, sum(1 for o in occurrences if o["recommended_delete"]))

    def test_find_ledger_duplicates_with_three_occurrences_mixed_completeness(self):
        # one bare row (clearly worse) plus two identical full-detail rows tied for best:
        # the bare one must still be recommended even though the top spot is a tie, and
        # among the tied pair one should also be recommended since they're identical.
        bare = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        full1 = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        full2 = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "c.csv"))
        result = self.app.merge_ledger_imports([bare, full1, full2])

        occurrences = self.app.find_ledger_duplicates(result["import_id"])[0]["occurrences"]
        self.assertEqual(3, len(occurrences))
        recommended = [o for o in occurrences if o["recommended_delete"]]
        kept = [o for o in occurrences if not o["recommended_delete"]]
        self.assertEqual(2, len(recommended))
        self.assertEqual(1, len(kept))
        self.assertTrue(all(o["data"]["ふりがな"] for o in kept))

    def test_find_ledger_duplicates_prefers_row_with_remarks(self):
        # rows identical in every field except 備考 (a note added during processing) —
        # the blank one should be recommended for deletion, keeping the annotated one.
        blank_note = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        with_note = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", "300*5400*21100*19"],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([blank_note, with_note])

        occurrences = self.app.find_ledger_duplicates(result["import_id"])[0]["occurrences"]
        recommended = next(o for o in occurrences if o["recommended_delete"])
        kept = next(o for o in occurrences if not o["recommended_delete"])
        self.assertEqual("", recommended["data"]["備考"])
        self.assertEqual("300*5400*21100*19", kept["data"]["備考"])

    def test_dismiss_ledger_duplicate_overrides_recommendation(self):
        # even a row the system recommends deleting can be kept: dismissing the group
        # removes it from the list entirely, regardless of any recommendation.
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])

        occurrences = self.app.find_ledger_duplicates(result["import_id"])[0]["occurrences"]
        self.assertTrue(any(o["recommended_delete"] for o in occurrences))
        self.app.dismiss_ledger_duplicate(result["import_id"], occurrences[0]["row_no"])

        self.assertEqual([], self.app.find_ledger_duplicates(result["import_id"]))
        # dismissing doesn't delete anything
        _, total = self.app.get_records(result["import_id"])
        self.assertEqual(2, total)

    def test_find_ledger_duplicates_no_recommendation_when_tied_but_different(self):
        # both rows tie on completeness AND actually differ in content (different
        # addresses) — genuinely ambiguous, so neither should be auto-flagged.
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所B", "000", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])

        occurrences = self.app.find_ledger_duplicates(result["import_id"])[0]["occurrences"]
        self.assertFalse(any(o["recommended_delete"] for o in occurrences))

    def test_delete_recommended_ledger_duplicates_bulk(self):
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
            ["2026-04-02", "B", "", "", "", "", "商品B", "1", "2000", "2000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
            ["2026-04-02", "B", "びー", "1991-01-01", "住所B", "000", "商品B", "1", "2000", "2000", ""],
        ], "b.csv"))
        result = self.app.merge_ledger_imports([first, second])
        self.assertEqual(2, len(result["duplicates"]))

        deleted = self.app.delete_recommended_ledger_duplicates(result["import_id"])

        self.assertEqual(2, deleted)
        self.assertEqual([], self.app.find_ledger_duplicates(result["import_id"]))
        _, total = self.app.get_records(result["import_id"])
        self.assertEqual(2, total)

    def test_delete_recommended_ledger_duplicates_bulk_crosses_chunk_boundary(self):
        # delete_recommended_ledger_duplicates batches its SQL in chunks of 500 targets
        # to stay well under SQLite's bound-parameter limit; use more than that to make
        # sure results spanning multiple chunks are still all correctly removed.
        n = 600
        bare_rows = [
            [f"2026-04-{(i % 27) + 1:02d}", f"Person{i}", "", "", "", "", f"商品{i}", "1", "1000", "1000", ""]
            for i in range(n)
        ]
        full_rows = [
            [f"2026-04-{(i % 27) + 1:02d}", f"Person{i}", "えー", "1990-01-01", "住所", "000", f"商品{i}", "1", "1000", "1000", ""]
            for i in range(n)
        ]
        first = self.app.import_ledger(self.write_ledger(bare_rows, "a.csv"))
        second = self.app.import_ledger(self.write_ledger(full_rows, "b.csv"))
        result = self.app.merge_ledger_imports([first, second])
        self.assertEqual(n, len(result["duplicates"]))

        deleted = self.app.delete_recommended_ledger_duplicates(result["import_id"])

        self.assertEqual(n, deleted)
        self.assertEqual([], self.app.find_ledger_duplicates(result["import_id"]))
        _, total = self.app.get_records(result["import_id"])
        self.assertEqual(n, total)

    def test_deleted_ledger_duplicate_does_not_reappear_after_re_merge(self):
        # auto_merge_ledger_imports() re-derives a brand-new merged import from ALL raw
        # imports every time (by design, to include newly-added files) — so deleting a
        # duplicate only from the merged copy would let it resurface on the next merge.
        # delete_ledger_record must also remove the matching row from its raw source.
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "えー", "1990-01-01", "住所A", "000", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        first_merge = self.app.merge_ledger_imports([first, second])
        occurrences = self.app.find_ledger_duplicates(first_merge["import_id"])[0]["occurrences"]
        bare_row = next(o for o in occurrences if not o["data"]["ふりがな"])
        self.app.delete_ledger_record(first_merge["import_id"], bare_row["row_no"])

        # a new, unrelated raw import triggers a fresh re-merge from ALL raw imports
        third = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "C", "しー", "1992-01-01", "住所C", "000", "商品C", "1", "3000", "3000", ""],
        ], "c.csv"))
        second_merge = self.app.merge_ledger_imports([first, second, third])

        self.assertEqual([], self.app.find_ledger_duplicates(second_merge["import_id"]))
        _, total = self.app.get_records(second_merge["import_id"])
        self.assertEqual(2, total)  # the kept duplicate row + the new unrelated row, not 3

    def test_dismissed_ledger_duplicate_stays_dismissed_after_re_merge(self):
        first = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "a.csv"))
        second = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "A", "", "", "", "", "商品A", "1", "1000", "1000", ""],
        ], "b.csv"))
        first_merge = self.app.merge_ledger_imports([first, second])
        occurrences = self.app.find_ledger_duplicates(first_merge["import_id"])[0]["occurrences"]
        self.app.dismiss_ledger_duplicate(first_merge["import_id"], occurrences[0]["row_no"])

        third = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "C", "しー", "1992-01-01", "住所C", "000", "商品C", "1", "3000", "3000", ""],
        ], "c.csv"))
        second_merge = self.app.merge_ledger_imports([first, second, third])

        self.assertEqual([], self.app.find_ledger_duplicates(second_merge["import_id"]))


class SuggestLedgerItemsTests(unittest.TestCase):
    def test_returns_none_without_api_key(self):
        with patch.dict(os.environ, {}, clear=True):
            self.assertIsNone(suggest_ledger_items(1000, {"A": 500}))

    def test_returns_none_with_empty_inventory(self):
        with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
            self.assertIsNone(suggest_ledger_items(1000, {}))

    @patch("openai.OpenAI")
    def test_filters_items_not_in_inventory_and_invalid_qty(self, mock_openai_cls):
        mock_client = MagicMock()
        mock_openai_cls.return_value = mock_client
        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps({
            "items": [
                {"product": "A", "qty": 2},
                {"product": "不明な商品", "qty": 1},
                {"product": "B", "qty": -1},
            ]
        })
        mock_client.chat.completions.create.return_value = mock_response
        with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
            result = suggest_ledger_items(1000, {"A": 500, "B": 300})
        self.assertEqual([{"product": "A", "qty": 2, "unit_cost": 500, "amount": 1000}], result)


class ToDateTests(unittest.TestCase):
    def test_accepts_hyphen_slash_and_dot_separators(self):
        from datetime import date
        expected = date(2025, 4, 23)
        self.assertEqual(expected, _to_date("2025-04-23"))
        self.assertEqual(expected, _to_date("2025/04/23"))
        self.assertEqual(expected, _to_date("2025.04.23"))
        self.assertEqual(expected, _to_date("2025.4.23"))


class FormatNumberTests(unittest.TestCase):
    def test_strips_trailing_zero_and_handles_float_rounding_error(self):
        from tax_system.web import format_number
        self.assertEqual(10800, format_number(10800.0))
        self.assertEqual(10800.5, format_number(10800.5))
        # Excel formula results can come back as e.g. 1980/1.1 == 1799.9999999999998
        self.assertEqual(1800, format_number(1799.9999999999998))
        self.assertEqual(485100, format_number(485100.00000000006))
        # non-float values (including numeric-looking strings like phone numbers) are untouched
        self.assertEqual("080-1234-5678", format_number("080-1234-5678"))
        self.assertEqual(3, format_number(3))
        self.assertIsNone(format_number(None))


class SafeUploadFilenameTests(unittest.TestCase):
    def test_preserves_extension_for_all_japanese_filename(self):
        # secure_filename() strips an entirely non-ASCII basename down to "", which
        # would otherwise swallow the extension along with it (Path("csv").suffix == "").
        from tax_system.web import safe_upload_filename
        result = safe_upload_filename("ワンピースカード　システム　買取伝票.csv", "upload-0")
        self.assertTrue(result.endswith(".csv"))
        self.assertEqual("upload-0.csv", result)

    def test_keeps_sanitized_ascii_basename(self):
        from tax_system.web import safe_upload_filename
        self.assertEqual("report.csv", safe_upload_filename("report.csv", "upload-0"))


class ComparisonLibraryTests(unittest.TestCase):
    HEADERS = ["年月日", "品目", "数量", "相手方名", "代価", "年月日", "数量"]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, rows, name="ledger.csv"):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_COLUMNS, row)))
        return path

    def write_export_data(self, rows, name="export_data.csv"):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=EXPORT_DATA_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(EXPORT_DATA_COLUMNS, row)))
        return path

    def register_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "輸出販売"
        for col, value in enumerate(self.HEADERS, 1):
            ws.cell(2, col).value = value
        path = self.root / "template.xlsx"
        wb.save(path)
        return self.app.register_template("comparison", path, "test", "v1", "2026-01-01")

    def build_one(self, ledger_rows, export_rows, ledger_name, export_name, template_id=None):
        self.app.import_ledger(self.write_ledger(ledger_rows, ledger_name))
        export_id = self.app.import_export_data(self.write_export_data(export_rows, export_name))
        template_id = template_id or self.register_template()
        return self.app.build_comparison(export_id, template_id)

    def test_month_summary_and_records_aggregate_across_imports(self):
        template_id = self.register_template()
        self.build_one(
            [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""]],
            [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]],
            "ledger1.csv", "export1.csv", template_id,
        )
        self.build_one(
            [["2026-01-02", "Bさん", "びー", "1991-01-01", "匿名住所", "000", "カードY", "1", "1200", "1200", ""]],
            [["2026-03-15", "カードY", "2200", "1", "2200", "海外顧客2", "銀行振込", "JPY"]],
            "ledger2.csv", "export2.csv", template_id,
        )

        summary = self.app.comparison_month_summary()
        self.assertEqual([{"month": "2026-03", "count": 2}], summary)

        rows, total = self.app.get_comparison_month_records("2026-03")
        self.assertEqual(2, total)
        self.assertEqual({"カードX", "カードY"}, {r["cells"][1] for r in rows})
        self.assertEqual(2, len({r["import_id"] for r in rows}))

    def test_find_comparison_duplicates_detects_identical_rows_across_imports(self):
        template_id = self.register_template()
        ledger_rows = [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "2", "1000", "2000", ""]]
        export_rows = [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]]
        older = self.build_one(ledger_rows, export_rows, "ledger1.csv", "export1.csv", template_id)
        newer = self.build_one(ledger_rows, export_rows, "ledger2.csv", "export2.csv", template_id)

        duplicates = self.app.find_comparison_duplicates()
        self.assertEqual(1, len(duplicates))
        occurrences = duplicates[0]["occurrences"]
        self.assertEqual(2, len(occurrences))
        by_import = {o["import_id"]: o for o in occurrences}
        self.assertTrue(by_import[older["import_id"]]["recommended_delete"])
        self.assertFalse(by_import[newer["import_id"]]["recommended_delete"])

    def test_find_comparison_duplicates_excludes_same_import_repeats(self):
        # two identical purchases + two identical sales, resolved within a single
        # build_comparison call: these repeat rows live in the SAME import and should
        # not be reported (same-file repeats are usually genuine separate transactions).
        template_id = self.register_template()
        ledger_rows = [
            ["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""],
            ["2026-01-02", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""],
        ]
        export_rows = [
            ["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"],
            ["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"],
        ]
        self.build_one(ledger_rows, export_rows, "ledger1.csv", "export1.csv", template_id)

        self.assertEqual([], self.app.find_comparison_duplicates())

    def test_find_comparison_duplicates_tolerates_float_rounding_noise(self):
        # two imports of what's really the same row, but a recalculated cell landed as
        # 999.9999999999999 in one file and 1000 in the other (Excel formula
        # re-evaluation) — these must still be recognized as the same duplicate.
        template_id = self.register_template()
        ledger_rows = [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""]]
        export_rows = [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]]
        self.build_one(ledger_rows, export_rows, "ledger1.csv", "export1.csv", template_id)
        second = self.build_one(ledger_rows, export_rows, "ledger2.csv", "export2.csv", template_id)

        with closing(self.app.connect()) as db, db:
            row = db.execute(
                "SELECT row_no, data_json FROM records WHERE import_id=?", (second["import_id"],)
            ).fetchone()
            data = json.loads(row["data_json"])
            data["values"] = [999.9999999999999 if v == 1000 else v for v in data["values"]]
            db.execute(
                "UPDATE records SET data_json=? WHERE import_id=? AND row_no=?",
                (json.dumps(data, ensure_ascii=False), second["import_id"], row["row_no"]),
            )

        duplicates = self.app.find_comparison_duplicates()
        self.assertEqual(1, len(duplicates))

    def test_delete_comparison_record_removes_row(self):
        template_id = self.register_template()
        ledger_rows = [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "2", "1000", "2000", ""]]
        export_rows = [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]]
        self.build_one(ledger_rows, export_rows, "ledger1.csv", "export1.csv", template_id)
        self.build_one(ledger_rows, export_rows, "ledger2.csv", "export2.csv", template_id)

        duplicates = self.app.find_comparison_duplicates()
        occ = next(o for o in duplicates[0]["occurrences"] if o["recommended_delete"])

        self.app.delete_comparison_record(occ["import_id"], occ["sheet"], occ["row_no"])

        self.assertEqual([], self.app.find_comparison_duplicates())
        _, total = self.app.get_comparison_month_records("2026-03")
        self.assertEqual(1, total)

    def test_delete_recommended_comparison_duplicates_bulk(self):
        template_id = self.register_template()
        # two independent duplicate pairs, each spanning two imports
        rows_x = (
            [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""]],
            [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]],
        )
        rows_y = (
            [["2026-01-01", "Bさん", "びー", "1990-01-01", "匿名住所", "000", "カードY", "1", "1500", "1500", ""]],
            [["2026-03-02", "カードY", "2500", "1", "2500", "海外顧客2", "銀行振込", "JPY"]],
        )
        self.build_one(rows_x[0], rows_x[1], "ledgerx1.csv", "exportx1.csv", template_id)
        self.build_one(rows_x[0], rows_x[1], "ledgerx2.csv", "exportx2.csv", template_id)
        self.build_one(rows_y[0], rows_y[1], "ledgery1.csv", "exporty1.csv", template_id)
        self.build_one(rows_y[0], rows_y[1], "ledgery2.csv", "exporty2.csv", template_id)

        self.assertEqual(2, len(self.app.find_comparison_duplicates()))

        deleted = self.app.delete_recommended_comparison_duplicates()

        self.assertEqual(2, deleted)
        self.assertEqual([], self.app.find_comparison_duplicates())
        _, total_x = self.app.get_comparison_month_records("2026-03")
        self.assertEqual(2, total_x)

    def test_export_comparison_month_combines_multiple_imports(self):
        template_id = self.register_template()
        self.build_one(
            [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""]],
            [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]],
            "ledger1.csv", "export1.csv", template_id,
        )
        self.build_one(
            [["2026-01-02", "Bさん", "びー", "1991-01-01", "匿名住所", "000", "カードY", "1", "1200", "1200", ""]],
            [["2026-03-15", "カードY", "2200", "1", "2200", "海外顧客2", "銀行振込", "JPY"]],
            "ledger2.csv", "export2.csv", template_id,
        )

        output = self.root / "export_2026-03.xlsx"
        result_path = self.app.export_comparison_month("2026-03", template_id, output)

        wb = load_workbook(result_path)
        ws = wb["輸出販売"]
        products = [ws.cell(r, 2).value for r in range(3, ws.max_row + 1) if ws.cell(r, 2).value]
        self.assertEqual({"カードX", "カードY"}, set(products))

    def test_export_comparison_month_raises_when_no_data(self):
        template_id = self.register_template()
        with self.assertRaises(ValueError):
            self.app.export_comparison_month("2099-01", template_id, self.root / "empty.xlsx")

    def test_export_comparison_month_requires_comparison_template(self):
        template_id = self.register_template()
        self.build_one(
            [["2026-01-01", "Aさん", "えー", "1990-01-01", "匿名住所", "000", "カードX", "1", "1000", "1000", ""]],
            [["2026-03-01", "カードX", "2000", "1", "2000", "海外顧客1", "銀行振込", "JPY"]],
            "ledger1.csv", "export1.csv", template_id,
        )
        with self.assertRaises(ValueError):
            self.app.export_comparison_month("2026-03", 99999, self.root / "bad.xlsx")


class FillComparisonPurchaseFromLedgerTests(unittest.TestCase):
    # matches the real production template: 品目 is a generic category shared by many
    # rows, 特徴 is the actual per-item description that lines up with 古物台帳's 商品名.
    HEADERS = [
        "年月日", "区別", "品目", "特徴", "Cert#", "単価(税込）", "数量", "代価（税込）", "種類", "相手方名", "備考",
        "年月日", "区分", "単価（JPY）", "数量", "代価（JPY）", "相手方名", "支払方法", "通貨",
    ]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, rows, name="ledger.csv"):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_COLUMNS, row)))
        return path

    def write_comparison_xlsx(self, rows, name="comparison.xlsx", sheet_name="輸出販売"):
        # rows: list of (region, product, feature, sale_date, sale_unit, sale_qty, sale_amount, buyer)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for col, value in enumerate(self.HEADERS, 1):
            ws.cell(2, col).value = value
        for r, (region, product, feature, sale_date, sale_unit, sale_qty, sale_amount, buyer) in enumerate(rows, 3):
            ws.cell(r, 2).value = region
            ws.cell(r, 3).value = product
            ws.cell(r, 4).value = feature
            ws.cell(r, 5).value = "ー"
            ws.cell(r, 12).value = sale_date
            ws.cell(r, 13).value = "売却（輸出）"
            ws.cell(r, 14).value = sale_unit
            ws.cell(r, 15).value = sale_qty
            ws.cell(r, 16).value = sale_amount
            ws.cell(r, 17).value = buyer
            ws.cell(r, 18).value = "クレジットカード"
            ws.cell(r, 19).value = "JPY"
        path = self.root / name
        wb.save(path)
        return path

    def test_fills_blank_purchase_side_from_matching_ledger_record(self):
        # fill_comparison_purchase_from_ledger reads only from the latest MERGED ledger
        # import (the deduped canonical set), so tests must merge before filling.
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "下置 誠龍", "しもおき", "1990-01-01", "住所", "000",
             "【傷あり特価】おれの時代だァ 【OP09-096】", "1", "5000", "5000", ""],
        ]))
        self.app.merge_ledger_imports([ledger_id])
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("買受", "ワンピースカード", "【傷あり特価】おれの時代だァ 【OP09-096】",
             "2026-06-03", 17800, 1, 17800, "Hani Fadlallah"),
        ]))

        result = self.app.fill_comparison_purchase_from_ledger()

        self.assertEqual(1, result["filled"])
        self.assertEqual(0, result["not_found"])
        records, _ = self.app.get_records(comparison_id)
        values = records[0]["data"]["values"]
        self.assertEqual("2026-05-01", str(values[0]))
        self.assertEqual(5000, values[5])  # 単価(税込)

    def test_fills_qty_amount_and_name_correctly(self):
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "下置 誠龍", "しもおき", "1990-01-01", "住所", "000",
             "特定カードA", "1", "5000", "5000", "元の備考"],
        ]))
        self.app.merge_ledger_imports([ledger_id])
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("買受", "ワンピースカード", "特定カードA", "2026-06-03", 17800, 1, 17800, "Hani Fadlallah"),
        ]))

        self.app.fill_comparison_purchase_from_ledger()

        records, _ = self.app.get_records(comparison_id)
        values = records[0]["data"]["values"]
        # purchase headers: 年月日,区別,品目,特徴,Cert#,単価(税込),数量,代価(税込),種類,相手方名,備考
        self.assertEqual(1, values[6])       # 数量
        self.assertEqual(5000, values[7])    # 代価（税込）
        self.assertEqual("下置 誠龍", values[9])  # 相手方名
        self.assertEqual("元の備考", values[10])  # 備考

    def test_does_not_reuse_a_ledger_record_across_two_sales(self):
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "客A", "きゃくえー", "1990-01-01", "住所A", "000", "特定カードB", "1", "1000", "1000", ""],
            ["2026-05-02", "客B", "きゃくびー", "1990-01-01", "住所B", "000", "特定カードB", "1", "1200", "1200", ""],
        ], "ledger.csv"))
        self.app.merge_ledger_imports([ledger_id])
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("買受", "カテゴリ", "特定カードB", "2026-06-01", 2000, 1, 2000, "海外客1"),
            ("買受", "カテゴリ", "特定カードB", "2026-06-02", 2200, 1, 2200, "海外客2"),
        ]))

        result = self.app.fill_comparison_purchase_from_ledger()

        self.assertEqual(2, result["filled"])
        records, _ = self.app.get_records(comparison_id)
        names = sorted(r["data"]["values"][9] for r in records)
        self.assertEqual(["客A", "客B"], names)  # FIFO: oldest ledger purchase used first

    def test_month_filter_scopes_to_sale_month(self):
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-04-01", "客A", "", "", "", "", "特定カードC", "1", "1000", "1000", ""],
        ]))
        self.app.merge_ledger_imports([ledger_id])
        self.app.import_comparison(self.write_comparison_xlsx([
            ("買受", "カテゴリ", "特定カードC", "2026-06-03", 2000, 1, 2000, "海外客"),
        ]))

        self.assertEqual(0, self.app.fill_comparison_purchase_from_ledger(month="2026-05")["filled"])
        self.assertEqual(1, self.app.fill_comparison_purchase_from_ledger(month="2026-06")["filled"])

    def test_not_found_when_no_matching_ledger_record(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            ("買受", "カテゴリ", "存在しないカード", "2026-06-03", 2000, 1, 2000, "海外客"),
        ]))

        result = self.app.fill_comparison_purchase_from_ledger()

        self.assertEqual(0, result["filled"])
        self.assertEqual(1, result["not_found"])

    def test_already_filled_purchase_rows_are_left_untouched(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "客A", "", "", "", "", "特定カードD", "1", "1000", "1000", ""],
        ]))
        path = self.write_comparison_xlsx([
            ("買受", "カテゴリ", "特定カードD", "2026-06-03", 2000, 1, 2000, "海外客"),
        ])
        wb = load_workbook(path)
        ws = wb.active
        ws.cell(3, 1).value = "2026-01-01"  # pre-fill purchase-side 年月日
        wb.save(path)
        comparison_id = self.app.import_comparison(path)

        result = self.app.fill_comparison_purchase_from_ledger()

        self.assertEqual(0, result["filled"])
        records, _ = self.app.get_records(comparison_id)
        self.assertNotEqual("客A", records[0]["data"]["values"][9])


class ComparisonPurchaseReuseTests(unittest.TestCase):
    # same real-world header layout as FillComparisonPurchaseFromLedgerTests, but the
    # rows here also fill in the purchase-side 年月日/相手方名 directly (as if the shop's
    # own process, or fill_comparison_purchase_from_ledger, had already populated them).
    HEADERS = [
        "年月日", "区別", "品目", "特徴", "Cert#", "単価(税込）", "数量", "代価（税込）", "種類", "相手方名", "備考",
        "年月日", "区分", "単価（JPY）", "数量", "代価（JPY）", "相手方名", "支払方法", "通貨",
    ]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_comparison_xlsx(self, rows, name="comparison.xlsx", sheet_name="輸出販売"):
        # rows: list of (purchase_date, feature, purchase_vendor, sale_date, sale_amount, buyer)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for col, value in enumerate(self.HEADERS, 1):
            ws.cell(2, col).value = value
        for r, (purchase_date, feature, vendor, sale_date, sale_amount, buyer) in enumerate(rows, 3):
            ws.cell(r, 1).value = purchase_date
            ws.cell(r, 2).value = "買受"
            ws.cell(r, 3).value = "カテゴリ"
            ws.cell(r, 4).value = feature
            ws.cell(r, 5).value = "ー"
            ws.cell(r, 10).value = vendor
            ws.cell(r, 12).value = sale_date
            ws.cell(r, 13).value = "売却（輸出）"
            ws.cell(r, 16).value = sale_amount
            ws.cell(r, 17).value = buyer
            ws.cell(r, 18).value = "クレジットカード"
            ws.cell(r, 19).value = "JPY"
        path = self.root / name
        wb.save(path)
        return path

    def test_detects_same_purchase_used_by_two_sales(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-03", 17800, "海外客1"),
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-10", 15000, "海外客2"),
        ]))

        results = self.app.find_comparison_purchase_reuse()

        self.assertEqual(1, len(results))
        self.assertEqual("下置 誠龍", results[0]["vendor"])
        self.assertEqual("特定カードX", results[0]["product"])
        self.assertEqual(2, len(results[0]["occurrences"]))

    def test_no_alert_when_purchases_are_distinct(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-03", 17800, "海外客1"),
            ("2026-05-02", "特定カードY", "下置 誠龍", "2026-06-10", 15000, "海外客2"),
        ]))

        self.assertEqual([], self.app.find_comparison_purchase_reuse())

    def test_detects_reuse_across_different_imports(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-03", 17800, "海外客1"),
        ], "comparison1.xlsx"))
        self.app.import_comparison(self.write_comparison_xlsx([
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-07-01", 16000, "海外客3"),
        ], "comparison2.xlsx"))

        results = self.app.find_comparison_purchase_reuse()

        self.assertEqual(1, len(results))
        self.assertEqual(2, len({o["import_id"] for o in results[0]["occurrences"]}))

    def test_ignores_rows_with_blank_purchase_side(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            (None, "特定カードX", None, "2026-06-03", 17800, "海外客1"),
            (None, "特定カードX", None, "2026-06-10", 15000, "海外客2"),
        ]))

        self.assertEqual([], self.app.find_comparison_purchase_reuse())

    def test_occurrences_include_sale_side_details_for_display(self):
        self.app.import_comparison(self.write_comparison_xlsx([
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-03", 17800, "海外客1"),
            ("2026-05-01", "特定カードX", "下置 誠龍", "2026-06-10", 15000, "海外客2"),
        ]))

        occurrences = self.app.find_comparison_purchase_reuse()[0]["occurrences"]
        by_buyer = {o["sale_buyer"]: o for o in occurrences}
        self.assertEqual("2026-06-03", by_buyer["海外客1"]["sale_date"])
        self.assertEqual(17800, by_buyer["海外客1"]["sale_amount"])
        self.assertEqual("2026-06-10", by_buyer["海外客2"]["sale_date"])
        self.assertEqual(15000, by_buyer["海外客2"]["sale_amount"])

    def test_records_page_for_row_finds_correct_page(self):
        rows = [
            ("2026-05-01", f"カード{i}", "下置 誠龍", "2026-06-03", 1000, "海外客")
            for i in range(150)
        ]
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx(rows))

        with closing(self.app.connect()) as db, db:
            row_nos = [r["row_no"] for r in db.execute(
                "SELECT row_no FROM records WHERE import_id=? ORDER BY row_no", (comparison_id,)
            ).fetchall()]

        self.assertEqual(1, self.app.records_page_for_row(comparison_id, "輸出販売", row_nos[0]))
        self.assertEqual(1, self.app.records_page_for_row(comparison_id, "輸出販売", row_nos[99]))
        self.assertEqual(2, self.app.records_page_for_row(comparison_id, "輸出販売", row_nos[100]))
        self.assertEqual(2, self.app.records_page_for_row(comparison_id, "輸出販売", row_nos[149]))


class LinkComparisonPurchaseManuallyTests(unittest.TestCase):
    # same real-world header layout used elsewhere: 品目=category, 特徴=specific item.
    HEADERS = [
        "年月日", "区別", "品目", "特徴", "Cert#", "単価(税込）", "数量", "代価（税込）", "種類", "相手方名", "備考",
        "年月日", "区分", "単価（JPY）", "数量", "代価（JPY）", "相手方名", "支払方法", "通貨",
    ]

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, rows, name="ledger.csv"):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(zip(LEDGER_COLUMNS, row)))
        return path

    def write_comparison_xlsx(self, rows, name="comparison.xlsx", sheet_name="輸出販売"):
        # rows: list of (feature, sale_date, sale_amount, buyer) — purchase side left blank
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for col, value in enumerate(self.HEADERS, 1):
            ws.cell(2, col).value = value
        for r, (feature, sale_date, sale_amount, buyer) in enumerate(rows, 3):
            ws.cell(r, 2).value = "買受"
            ws.cell(r, 3).value = "未開封BOXより"
            ws.cell(r, 4).value = feature
            ws.cell(r, 5).value = "ー"
            ws.cell(r, 12).value = sale_date
            ws.cell(r, 13).value = "売却（輸出）"
            ws.cell(r, 16).value = sale_amount
            ws.cell(r, 17).value = buyer
            ws.cell(r, 18).value = "クレジットカード"
            ws.cell(r, 19).value = "JPY"
        path = self.root / name
        wb.save(path)
        return path

    def test_fills_purchase_side_with_manual_prorated_amount(self):
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "箱売り太郎", "はこうりたろう", "1990-01-01", "住所", "000",
             "未開封BOX ワンピースカード OP-01", "1", "50000", "50000", ""],
        ]))
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("モンキー・D・ルフィ 【SR】【パラレル】", "2026-06-03", 3000, "海外客1"),
        ]))
        records, _ = self.app.get_records(ledger_id)
        ledger_record_id = records[0]["id"]

        self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", 3, ledger_record_id, qty=1, amount=500)

        records, _ = self.app.get_records(comparison_id)
        values = records[0]["data"]["values"]
        self.assertEqual("2026-05-01", str(values[0]))  # 年月日
        self.assertEqual(1, values[6])                  # 数量
        self.assertEqual(500, values[5])                 # 単価(税込)
        self.assertEqual(500, values[7])                 # 代価（税込）
        self.assertEqual("箱売り太郎", values[9])         # 相手方名
        self.assertIn("未開封BOX", values[10])            # 備考

    def test_search_ledger_caps_results_and_reports_total_match_count(self):
        # a common card can have far more than 30 unconsumed purchase records; the total
        # must still be reported so the caller knows results were truncated, not exhaustive.
        rows = [
            [f"2026-01-{d:02d}", f"客{d}", "", "", "", "", "同じカード", "1", "1000", "1000", ""]
            for d in range(1, 32)
        ]
        self.app.import_ledger(self.write_ledger(rows))

        results, total = self.app.search_ledger("同じカード")

        self.assertEqual(31, total)
        self.assertEqual(30, len(results))

    def test_search_ledger_orders_oldest_purchase_first(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-03-01", "客A", "", "", "", "", "同じカード", "1", "1000", "1000", ""],
            ["2026-01-01", "客B", "", "", "", "", "同じカード", "1", "1000", "1000", ""],
            ["2026-02-01", "客C", "", "", "", "", "同じカード", "1", "1000", "1000", ""],
        ]))

        results, total = self.app.search_ledger("同じカード")

        self.assertEqual(3, total)
        self.assertEqual(["客B", "客C", "客A"], [r["name"] for r in results])

    def test_search_ledger_can_narrow_to_a_purchase_month(self):
        self.app.import_ledger(self.write_ledger([
            ["2026-01-01", "客A", "", "", "", "", "同じカード", "1", "1000", "1000", ""],
            ["2026-06-01", "客B", "", "", "", "", "同じカード", "1", "1000", "1000", ""],
        ]))

        results, total = self.app.search_ledger("同じカード", month="2026-06")

        self.assertEqual(1, total)
        self.assertEqual("客B", results[0]["name"])

    def test_ledger_record_stays_available_for_reuse_on_another_row(self):
        # one unopened box can be prorated across many individually-sold cards, so linking
        # it once must not mark it "consumed" the way fill_comparison_purchase_from_ledger does.
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "箱売り太郎", "はこうりたろう", "1990-01-01", "住所", "000",
             "未開封BOX ワンピースカード OP-01", "1", "50000", "50000", ""],
        ]))
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("モンキー・D・ルフィ 【SR】【パラレル】", "2026-06-03", 3000, "海外客1"),
            ("ナミ 【SR】【パラレル】", "2026-06-04", 2000, "海外客2"),
        ]))
        records, _ = self.app.get_records(ledger_id)
        ledger_record_id = records[0]["id"]

        self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", 3, ledger_record_id, qty=1, amount=500)
        # still findable/usable for a second row from the same box
        results, _ = self.app.search_ledger("未開封BOX")
        self.assertTrue(any(c["id"] == ledger_record_id for c in results))
        self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", 4, ledger_record_id, qty=1, amount=300)

        records, _ = self.app.get_records(comparison_id)
        by_row = {r["row_no"]: r["data"]["values"] for r in records}
        self.assertEqual(500, by_row[3][7])
        self.assertEqual(300, by_row[4][7])

    def test_note_is_generic_when_ledger_product_name_is_blank(self):
        # 明細未作成 (a purchase header with no 商品名 yet, e.g. an unopened box that hasn't
        # been itemized) must not produce an awkward empty-quotes note like 「」.
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "箱売り太郎", "はこうりたろう", "1990-01-01", "住所", "000", "", "1", "50000", "50000", ""],
        ]))
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("モンキー・D・ルフィ 【SR】【パラレル】", "2026-06-03", 3000, "海外客1"),
        ]))
        records, _ = self.app.get_records(ledger_id)
        ledger_record_id = records[0]["id"]

        self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", 3, ledger_record_id, qty=1, amount=500)

        records, _ = self.app.get_records(comparison_id)
        note = records[0]["data"]["values"][10]
        self.assertNotIn("「」", note)
        self.assertEqual("内訳未確定の仕入から案分", note)

    def test_raises_when_row_not_found(self):
        ledger_id = self.app.import_ledger(self.write_ledger([
            ["2026-05-01", "箱売り太郎", "", "", "", "", "未開封BOX", "1", "50000", "50000", ""],
        ]))
        records, _ = self.app.get_records(ledger_id)
        with self.assertRaises(ValueError):
            self.app.link_comparison_purchase_manually(99999, "輸出販売", 3, records[0]["id"], qty=1, amount=500)

    def test_raises_when_ledger_record_not_found(self):
        comparison_id = self.app.import_comparison(self.write_comparison_xlsx([
            ("モンキー・D・ルフィ 【SR】【パラレル】", "2026-06-03", 3000, "海外客1"),
        ]))
        with self.assertRaises(ValueError):
            self.app.link_comparison_purchase_manually(comparison_id, "輸出販売", 3, 99999, qty=1, amount=500)


class DeleteImportsTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_ledger(self, name):
        path = self.root / name
        row = dict(zip(LEDGER_COLUMNS, [
            "2026-01-01", "テスト", "てすと", "2000-01-01", "匿名住所", "000", "商品", "1", "100", "100", "",
        ]))
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n")
            writer.writeheader(); writer.writerow(row)
        return path

    def test_delete_imports_removes_selected_and_skips_missing(self):
        id1 = self.app.import_ledger(self.write_ledger("l1.csv"))
        id2 = self.app.import_ledger(self.write_ledger("l2.csv"))
        id3 = self.app.import_ledger(self.write_ledger("l3.csv"))

        count = self.app.delete_imports([id1, id2, 99999])

        self.assertEqual(2, count)
        self.assertIsNone(self.app.get_import(id1))
        self.assertIsNone(self.app.get_import(id2))
        self.assertIsNotNone(self.app.get_import(id3))


class InventoryExportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.app = TaxSystem(self.root / "runtime")

    def tearDown(self):
        self.tmp.cleanup()

    def write_inventory(self, name="inventory.csv"):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=INVENTORY_COLUMNS, lineterminator="\r\n")
            writer.writeheader()
            writer.writerow({"商品名": "テストカードA", "仕入れ原価": "5000", "在庫数": "10"})
            writer.writerow({"商品名": "テストカードB", "仕入れ原価": "3000", "在庫数": "5"})
        return path

    def test_export_inventory_writes_csv_with_display_columns(self):
        import_id = self.app.import_inventory(self.write_inventory(), as_of="2026-04")

        output = self.root / "out.csv"
        self.app.export(import_id, output, preview=True)

        with output.open(encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
        self.assertEqual(["商品名", "カテゴリ", "サブカテゴリ", "グループ", "仕入れ原価", "販売価格", "在庫数"], list(rows[0].keys()))
        self.assertEqual("テストカードA", rows[0]["商品名"])
        self.assertEqual("5000", rows[0]["仕入れ原価"])
        self.assertEqual("", rows[0]["カテゴリ"])

    def test_export_inventory_month_requires_matching_as_of(self):
        import_id = self.app.import_inventory(self.write_inventory(), as_of="2026-04")

        self.app.export(import_id, self.root / "ok.csv", preview=True, month="2026-04")

        with self.assertRaises(ValueError):
            self.app.export(import_id, self.root / "bad.csv", preview=True, month="2026-05")


if __name__ == "__main__": unittest.main()

