from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import tempfile
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


LEDGER_COLUMNS = ["日時", "名前", "ふりがな", "生年月日", "住所", "電話番号", "商品名", "個数", "単価", "金額", "備考"]
INVENTORY_COLUMNS = ["商品名", "仕入れ原価", "在庫数"]
EXPORT_DATA_COLUMNS = ["年月日", "品名", "金額", "数量", "小計", "相手方名", "支払方法", "通貨"]
LEDGER_POS_COLUMNS = ["履歴ID", "状態", "日時", "ユーザーID", "氏名", "カード番号", "カード名", "数量", "単価", "金額", "カード備考", "全体備考"]
LEDGER_IDENTITY_COLUMNS = ["日時", "名前", "名前ふりがな", "生年月日", "住所", "電話番号", "金額"]


def translate_ja_to_en(text: str) -> str | None:
    """商品名をインボイス用の簡潔な英語名に翻訳する。OPENAI_API_KEY が未設定、
    またはAPI呼び出しに失敗した場合は None を返す（呼び出し側で手入力に切り替える）。
    """
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key or not text.strip():
        return None
    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0,
            messages=[
                {"role": "system", "content": (
                    "あなたは古物・トレーディングカード商品の名称を、輸出インボイスに記載する"
                    "簡潔な英語名に翻訳するアシスタントです。説明や注釈を付けず、英語名だけを"
                    "1行で答えてください。"
                )},
                {"role": "user", "content": text},
            ],
        )
        result = (response.choices[0].message.content or "").strip()
        return result or None
    except Exception:
        return None


def suggest_ledger_items(remainder: float, inventory: dict[str, float]) -> list[dict[str, Any]] | None:
    """内訳復元の残額と期末在庫表の商品リストから、残額に合う商品の組み合わせをAIに提案してもらう。
    提案はあくまで参考で、実際に追加するかどうかは呼び出し側（人）が選ぶ。
    OPENAI_API_KEY未設定・在庫が空・API呼び出し失敗時はNoneを返す（呼び出し側は手動検索に切り替える）。
    """
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key or not inventory:
        return None
    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)
        catalog = "\n".join(f"- {name}: 仕入れ原価 {cost}円" for name, cost in inventory.items())
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": (
                    "あなたは古物台帳の内訳復元を手伝うアシスタントです。指定された商品リストの中から、"
                    "合計金額ができるだけ指定の残額に一致する組み合わせを選んでください。"
                    'JSON形式 {"items": [{"product": "商品名", "qty": 数量}, ...]} のみで回答してください。'
                    "リストにない商品名は使わないでください。ぴったり一致する組み合わせがなければ、"
                    "最も近いものを1つ提案してください。"
                )},
                {"role": "user", "content": f"残額: {remainder}円\n商品リスト:\n{catalog}"},
            ],
        )
        payload = json.loads(response.choices[0].message.content or "{}")
        items = payload.get("items")
        if not isinstance(items, list):
            return None
        result = []
        for item in items:
            product = item.get("product") if isinstance(item, dict) else None
            qty = item.get("qty") if isinstance(item, dict) else None
            if product in inventory and isinstance(qty, (int, float)) and qty > 0:
                cost = inventory[product]
                result.append({"product": product, "qty": qty, "unit_cost": cost, "amount": cost * qty})
        return result or None
    except Exception:
        return None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    match = re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", str(value).strip())
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _read_csv_text(source: Path) -> tuple[str, str]:
    if source.suffix.lower() != ".csv":
        raise ValueError(
            f"CSVファイルを選択してください（{source.name} は .csv ではないようです。"
            "Excelファイルの場合は取込画面で正しい種類を選んでください）"
        )
    raw = source.read_bytes()
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError(f"{source.name} の文字コードを判別できません（UTF-8・Shift-JISのいずれでもありません）")


def _is_header_echo(values: Any, columns: list[str]) -> bool:
    """複数回のエクスポートを1つのファイルに貼り合わせた際、ヘッダー行が
    そのままデータ行として紛れ込むことがある（値が列名と完全に一致する行）。
    そのようなゴミ行を検出する。
    """
    return list(values) == columns


def _index_of(headers: list[Any], name: str, contains: bool = False) -> int | None:
    for index, header in enumerate(headers):
        if header is None:
            continue
        if contains and name in str(header):
            return index
        if not contains and header == name:
            return index
    return None


@dataclass
class CheckResult:
    code: str
    level: str
    message: str
    sheet: str | None = None
    row_no: int | None = None


class TaxSystem:
    def __init__(self, home: str | Path = ".tax-system") -> None:
        self.home = Path(home).resolve()
        self.templates = self.home / "templates"
        self.db_path = self.home / "tax-system.sqlite3"

    def initialize(self) -> None:
        self.templates.mkdir(parents=True, exist_ok=True)
        with closing(self.connect()) as db, db:
            db.executescript(
                """
                PRAGMA journal_mode=WAL;
                CREATE TABLE IF NOT EXISTS template_versions(
                  id INTEGER PRIMARY KEY, report_type TEXT NOT NULL, name TEXT NOT NULL,
                  version TEXT NOT NULL, effective_from TEXT NOT NULL, format TEXT NOT NULL,
                  source_name TEXT NOT NULL, stored_path TEXT NOT NULL, sha256 TEXT NOT NULL,
                  settings_json TEXT NOT NULL, created_at TEXT NOT NULL,
                  UNIQUE(report_type, name, version)
                );
                CREATE TABLE IF NOT EXISTS imports(
                  id INTEGER PRIMARY KEY, kind TEXT NOT NULL, source_name TEXT NOT NULL,
                  source_sha256 TEXT NOT NULL, imported_at TEXT NOT NULL,
                  metadata_json TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS records(
                  id INTEGER PRIMARY KEY, import_id INTEGER NOT NULL REFERENCES imports(id),
                  sheet_name TEXT, row_no INTEGER NOT NULL, data_json TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS exports(
                  id INTEGER PRIMARY KEY, import_id INTEGER NOT NULL, template_id INTEGER,
                  mode TEXT NOT NULL, output_name TEXT NOT NULL, output_sha256 TEXT NOT NULL,
                  checks_json TEXT NOT NULL, created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS allocations(
                  id INTEGER PRIMARY KEY, sale_import_id INTEGER NOT NULL, sale_sheet TEXT NOT NULL,
                  sale_row_no INTEGER NOT NULL, ledger_record_id INTEGER, status TEXT NOT NULL,
                  candidates_json TEXT NOT NULL, note TEXT, created_at TEXT NOT NULL,
                  UNIQUE(sale_import_id, sale_sheet, sale_row_no)
                );
                CREATE TABLE IF NOT EXISTS ledger_items(
                  id INTEGER PRIMARY KEY, ledger_import_id INTEGER NOT NULL, ledger_row_no INTEGER NOT NULL,
                  product TEXT NOT NULL, qty REAL NOT NULL, unit_cost REAL NOT NULL, amount REAL NOT NULL,
                  source TEXT NOT NULL, created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS ledger_duplicate_dismissals(
                  import_id INTEGER NOT NULL, key_json TEXT NOT NULL, created_at TEXT NOT NULL,
                  PRIMARY KEY(import_id, key_json)
                );
                """
            )

    def connect(self) -> sqlite3.Connection:
        self.home.mkdir(parents=True, exist_ok=True)
        db = sqlite3.connect(self.db_path)
        db.row_factory = sqlite3.Row
        return db

    def register_template(self, report_type: str, source: str | Path, name: str,
                          version: str, effective_from: str) -> int:
        self.initialize()
        source = Path(source).resolve(strict=True)
        suffix = source.suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            raise ValueError("テンプレートは .xlsx または .csv のみ登録できます")
        target_dir = self.templates / report_type / name / version
        target_dir.mkdir(parents=True, exist_ok=False)
        target = target_dir / ("template" + suffix)
        shutil.copy2(source, target)
        settings = self._template_settings(target)
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self.connect()) as db, db:
            cur = db.execute(
                "INSERT INTO template_versions(report_type,name,version,effective_from,format,source_name,stored_path,sha256,settings_json,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (report_type, name, version, effective_from, suffix[1:], source.name,
                 str(target), sha256(target), json_text(settings), now),
            )
            return int(cur.lastrowid)

    def _template_settings(self, path: Path) -> dict[str, Any]:
        if path.suffix.lower() == ".csv":
            raw = path.read_bytes()
            encoding = "utf-8-sig" if raw.startswith(b"\xef\xbb\xbf") else "utf-8"
            try:
                text = raw.decode(encoding)
            except UnicodeDecodeError:
                encoding, text = "cp932", raw.decode("cp932")
            dialect = csv.Sniffer().sniff(text[:8192])
            rows = list(csv.reader(text.splitlines(), dialect))
            return {"encoding": encoding, "bom": raw.startswith(b"\xef\xbb\xbf"),
                    "delimiter": dialect.delimiter, "lineterminator": "CRLF" if b"\r\n" in raw else "LF",
                    "header": rows[0] if rows else []}
        wb = load_workbook(path, read_only=False, data_only=False)
        return {"sheets": [{"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column,
                            "headers": [ws.cell(2, col).value for col in range(1, ws.max_column + 1)],
                            "merged": [str(r) for r in ws.merged_cells.ranges]}
                           for ws in wb.worksheets]}

    def import_ledger(self, source: str | Path) -> int:
        self.initialize()
        source = Path(source).resolve(strict=True)
        text, encoding = _read_csv_text(source)
        reader = csv.DictReader(text.splitlines())
        if reader.fieldnames != LEDGER_COLUMNS:
            raise ValueError(f"古物台帳の列が既存仕様と一致しません: {reader.fieldnames}")
        rows = [row for row in reader if not _is_header_echo(row.values(), LEDGER_COLUMNS)]
        return self._save_import("ledger", source.name, sha256(source), rows, {"encoding": encoding, "columns": reader.fieldnames})

    def import_ledger_pos(self, source: str | Path) -> dict[str, Any]:
        """POSの取引CSV（履歴ID/状態/日時/ユーザーID/氏名/カード番号/カード名/数量/単価/金額/カード備考/全体備考）を
        古物台帳の標準形式に変換して取り込む。本人確認情報（ふりがな・生年月日・住所・電話番号）は
        この形式には無いため空欄のまま登録する。「状態」が「承認済み」以外の行は対象外とし、件数を報告する。
        """
        self.initialize()
        source = Path(source).resolve(strict=True)
        text, encoding = _read_csv_text(source)
        reader = csv.DictReader(text.splitlines())
        if reader.fieldnames != LEDGER_POS_COLUMNS:
            raise ValueError(f"POS取引データの列が想定と一致しません（{LEDGER_POS_COLUMNS}）: {reader.fieldnames}")
        rows: list[dict[str, Any]] = []
        skipped = 0
        for row in reader:
            if row.get("状態") != "承認済み":
                skipped += 1
                continue
            note = " ".join(part for part in (row.get("カード備考"), row.get("全体備考")) if part).strip()
            rows.append({k: "" for k in LEDGER_COLUMNS} | {
                "日時": row.get("日時", ""), "名前": row.get("氏名", ""),
                "商品名": row.get("カード名", ""), "個数": row.get("数量", ""),
                "単価": row.get("単価", ""), "金額": row.get("金額", ""), "備考": note,
            })
        import_id = self._save_import(
            "ledger", source.name, sha256(source), rows,
            {"encoding": encoding, "source_format": "pos_csv", "skipped_not_approved": skipped},
        )
        return {"import_id": import_id, "imported": len(rows), "skipped": skipped}

    def _build_ledger_identity_rows(self, header: list[Any], raw_rows: Iterable[Any]) -> list[dict[str, Any]]:
        # メールアドレス列は取込先の古物台帳の形式に無いため、あっても無視する（無くても可）。
        with_email = LEDGER_IDENTITY_COLUMNS[:-1] + ["メールアドレス", LEDGER_IDENTITY_COLUMNS[-1]]
        if header != LEDGER_IDENTITY_COLUMNS and header != with_email:
            raise ValueError(f"本人確認データの列が想定と一致しません（{LEDGER_IDENTITY_COLUMNS}）: {header}")
        rows: list[dict[str, Any]] = []
        for values in raw_rows:
            values = list(values)[: len(header)]
            if all(v in (None, "") for v in values) or _is_header_echo(values, header):
                continue
            data = dict(zip(header, values))
            rows.append({k: "" for k in LEDGER_COLUMNS} | {
                "日時": data.get("日時") or "", "名前": data.get("名前") or "",
                "ふりがな": data.get("名前ふりがな") or "", "生年月日": data.get("生年月日") or "",
                "住所": data.get("住所") or "", "電話番号": data.get("電話番号") or "",
                "金額": data.get("金額") if data.get("金額") is not None else "",
            })
        return rows

    def import_ledger_identity(self, source: str | Path) -> dict[str, Any]:
        """本人確認データ（日時/名前/名前ふりがな/生年月日/住所/電話番号/[メールアドレス]/金額）を
        古物台帳の標準形式に変換して取り込む。商品情報（商品名・個数・単価）と備考は
        この形式には無いため空欄のまま登録する。メールアドレス列があっても無視する。
        Excel（.xlsx）とCSVの両方に対応する。
        """
        self.initialize()
        source = Path(source).resolve(strict=True)
        suffix = source.suffix.lower()
        metadata: dict[str, Any] = {}
        if suffix == ".xlsx":
            wb = load_workbook(source, read_only=True, data_only=True)
            try:
                ws = wb.worksheets[0]
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                while header and header[-1] is None:
                    header.pop()
                raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
            finally:
                wb.close()
            metadata["source_format"] = "identity_xlsx"
        elif suffix == ".csv":
            text, encoding = _read_csv_text(source)
            csv_rows = list(csv.reader(text.splitlines()))
            header = csv_rows[0] if csv_rows else []
            raw_rows = csv_rows[1:]
            metadata["source_format"] = "identity_csv"
            metadata["encoding"] = encoding
        else:
            raise ValueError(f"Excel（.xlsx）またはCSVファイルを選択してください: {source.name}")

        rows = self._build_ledger_identity_rows(header, raw_rows)
        import_id = self._save_import("ledger", source.name, sha256(source), rows, metadata)
        return {"import_id": import_id, "imported": len(rows)}

    def import_comparison(self, source: str | Path) -> int:
        self.initialize()
        source = Path(source).resolve(strict=True)
        if source.suffix.lower() != ".xlsx":
            raise ValueError(f"Excelファイル（.xlsx）を選択してください: {source.name}")
        # data_only=True: read the cached calculated value for formula cells (e.g. 代価=単価×数量)
        # rather than the formula text itself, so validation/matching see real numbers and the
        # records view shows the actual amount instead of "=F3*G3".
        wb = load_workbook(source, read_only=False, data_only=True)
        records: list[dict[str, Any]] = []
        sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
            if split is None:
                raise ValueError(f"{ws.title}: 受入れ・払出しの境界を判定できません")
            sheets.append({"name": ws.title, "headers": headers, "max_column": ws.max_column})
            for row_no in range(3, ws.max_row + 1):
                cells = [ws.cell(row_no, c) for c in range(1, ws.max_column + 1)]
                values = [c.value for c in cells]
                # Existing sheets contain formulas in otherwise unused rows. Dates are
                # the authoritative indication that a purchase/sale row is populated.
                if values[0] in (None, "") and values[split] in (None, ""):
                    continue
                records.append({"sheet": ws.title, "row_no": row_no, "values": values})
        return self._save_import("comparison", source.name, sha256(source), records, {"sheets": sheets})

    def import_inventory(self, source: str | Path, as_of: str | None = None) -> int:
        self.initialize()
        source = Path(source).resolve(strict=True)
        text, encoding = _read_csv_text(source)
        reader = csv.DictReader(text.splitlines())
        if reader.fieldnames != INVENTORY_COLUMNS:
            raise ValueError(f"期末在庫表の列が想定と一致しません（{INVENTORY_COLUMNS}）: {reader.fieldnames}")
        rows = list(reader)
        return self._save_import(
            "inventory", source.name, sha256(source), rows,
            {"encoding": encoding, "columns": reader.fieldnames, "as_of": as_of},
        )

    def import_inventory_products(self, source: str | Path, as_of: str | None = None) -> dict[str, Any]:
        """ECサイトの商品CSV（商品番号・写真・SEO設定など多数の列を含む商品マスタ）から、
        商品名・仕入・在庫数の3列だけを抜き出して期末在庫表として取り込む。
        as_of（例: "2026-04"）は、この在庫表がいつ時点のものかを示す基準年月。
        """
        self.initialize()
        source = Path(source).resolve(strict=True)
        text, encoding = _read_csv_text(source)
        reader = csv.DictReader(text.splitlines())
        required = {"商品名", "在庫数", "仕入"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise ValueError(f"商品CSVの列に「商品名」「在庫数」「仕入」が見つかりません: {reader.fieldnames}")
        rows: list[dict[str, Any]] = []
        for row in reader:
            product = (row.get("商品名") or "").strip()
            if not product:
                continue
            rows.append({
                "商品名": product, "仕入れ原価": row.get("仕入", ""), "在庫数": row.get("在庫数", ""),
                "カテゴリ": row.get("カテゴリ", ""), "サブカテゴリ": row.get("サブカテゴリ", ""),
                "グループ": row.get("グループ", ""), "販売価格": row.get("販売価格", ""),
            })
        import_id = self._save_import(
            "inventory", source.name, sha256(source), rows,
            {"encoding": encoding, "source_format": "ec_products_csv", "as_of": as_of},
        )
        return {"import_id": import_id, "imported": len(rows)}

    def import_export_data(self, source: str | Path) -> int:
        self.initialize()
        source = Path(source).resolve(strict=True)
        text, encoding = _read_csv_text(source)
        reader = csv.DictReader(text.splitlines())
        if reader.fieldnames != EXPORT_DATA_COLUMNS:
            raise ValueError(f"輸出データの列が想定と一致しません（{EXPORT_DATA_COLUMNS}）: {reader.fieldnames}")
        rows = list(reader)
        return self._save_import("export_data", source.name, sha256(source), rows, {"encoding": encoding, "columns": reader.fieldnames})

    def import_auto(self, source: str | Path, inventory_as_of: str | None = None) -> dict[str, Any]:
        """ファイルの拡張子と列構成から種類を自動判定して取り込む（一括取込用）。
        各形式の列は互いに重ならないため、既存の取込処理を順番に試し、
        列が一致したものをそのまま採用する。inventory_as_of は期末在庫表と判定された
        場合にのみ使われる基準年月（例: "2026-04"）。
        """
        source = Path(source).resolve(strict=True)
        suffix = source.suffix.lower()
        if suffix == ".csv":
            candidates: list[tuple[str, Any]] = [
                ("古物台帳", lambda: {"kind": "ledger", "import_id": self.import_ledger(source)}),
                ("古物台帳（POS取引データ）", lambda: {"kind": "ledger", **self.import_ledger_pos(source)}),
                ("古物台帳（本人確認データ）", lambda: {"kind": "ledger", **self.import_ledger_identity(source)}),
                ("期末在庫表", lambda: {"kind": "inventory", "import_id": self.import_inventory(source, inventory_as_of)}),
                ("期末在庫表（ECサイト商品CSV）", lambda: {"kind": "inventory", **self.import_inventory_products(source, inventory_as_of)}),
                ("輸出データ", lambda: {"kind": "export_data", "import_id": self.import_export_data(source)}),
            ]
        elif suffix == ".xlsx":
            candidates = [
                ("相対表", lambda: {"kind": "comparison", "import_id": self.import_comparison(source)}),
                ("古物台帳（本人確認データ）", lambda: {"kind": "ledger", **self.import_ledger_identity(source)}),
            ]
        else:
            raise ValueError(f"対応していないファイル形式です: {source.name}")

        for label, fn in candidates:
            try:
                result = fn()
            except ValueError:
                continue
            return {"label": label, **result}
        known = "・".join(label for label, _ in candidates)
        raise ValueError(f"列構成から種類を判別できませんでした（{known} のいずれでもありません）")

    def record_export_entry(self, rows: list[dict[str, Any]], label: str) -> int:
        """画面から手入力した輸出取引（品名・数量・単価などを1件ずつ追加したもの）を
        輸出データの取込として保存する。ファイルではなく手入力のため、参照用のラベルのみ記録する。
        """
        self.initialize()
        return self._save_import(
            "export_data", label, "manual-entry", rows,
            {"encoding": "manual", "source_format": "manual_entry", "columns": EXPORT_DATA_COLUMNS},
        )

    def _save_import(self, kind: str, source_name: str, source_hash: str,
                     rows: Iterable[dict[str, Any]], metadata: dict[str, Any]) -> int:
        now = datetime.now().isoformat(timespec="seconds")
        rows = list(rows)
        with closing(self.connect()) as db, db:
            cur = db.execute("INSERT INTO imports(kind,source_name,source_sha256,imported_at,metadata_json) VALUES(?,?,?,?,?)",
                             (kind, source_name, source_hash, now, json_text(metadata)))
            import_id = int(cur.lastrowid)
            payload = []
            for index, row in enumerate(rows, 2):
                sheet = row.get("sheet")
                row_no = int(row.get("row_no", index))
                payload.append((import_id, sheet, row_no, json_text(row)))
            db.executemany("INSERT INTO records(import_id,sheet_name,row_no,data_json) VALUES(?,?,?,?)", payload)
        return import_id

    def list_templates(self) -> list[dict[str, Any]]:
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT id, report_type, name, version, effective_from, format, created_at FROM template_versions ORDER BY report_type, name, effective_from DESC"
            ).fetchall()
        return [dict(row) for row in rows]

    def list_imports(self) -> list[dict[str, Any]]:
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT id, kind, source_name, imported_at, metadata_json FROM imports ORDER BY id DESC"
            ).fetchall()
        results = []
        for row in rows:
            item = dict(row)
            item["as_of"] = json.loads(item.pop("metadata_json")).get("as_of")
            results.append(item)
        return results

    def get_import(self, import_id: int) -> dict[str, Any] | None:
        with closing(self.connect()) as db, db:
            row = db.execute(
                "SELECT id, kind, source_name, imported_at, metadata_json FROM imports WHERE id=?", (import_id,)
            ).fetchone()
        if not row:
            return None
        data = dict(row)
        data["metadata"] = json.loads(data.pop("metadata_json"))
        return data

    def _record_month(self, kind: str, data: dict[str, Any], sheet_headers: dict[str, list] | None) -> str | None:
        if kind == "ledger":
            d = _to_date(data.get("日時"))
        elif kind == "export_data":
            d = _to_date(data.get("年月日"))
        elif kind == "comparison" and sheet_headers:
            headers = sheet_headers.get(data.get("sheet"))
            if not headers:
                return None
            split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
            values = data.get("values", [])
            d = _to_date(values[split]) if split is not None and split < len(values) else None
        else:
            return None
        return f"{d.year:04d}-{d.month:02d}" if d else None

    def get_records(self, import_id: int, sheet: str | None = None, month: str | None = None,
                    sort: str | None = None, sort_dir: str = "asc",
                    offset: int = 0, limit: int = 100) -> tuple[list[dict[str, Any]], int]:
        where = "import_id=?"
        params: list[Any] = [import_id]
        if sheet is not None:
            where += " AND sheet_name=?"
            params.append(sheet)
        if month is None and sort is None:
            with closing(self.connect()) as db, db:
                total = db.execute(f"SELECT COUNT(*) FROM records WHERE {where}", params).fetchone()[0]
                rows = db.execute(
                    f"SELECT * FROM records WHERE {where} ORDER BY sheet_name, row_no LIMIT ? OFFSET ?",
                    params + [limit, offset],
                ).fetchall()
            results = []
            for row in rows:
                item = dict(row)
                item["data"] = json.loads(item.pop("data_json"))
                results.append(item)
            return results, total

        # month filtering / sorting inspect the JSON data, so they can't be pushed into SQL;
        # fetch everything for the import and filter/sort/paginate in Python instead.
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT kind, metadata_json FROM imports WHERE id=?", (import_id,)).fetchone()
            rows = db.execute(
                f"SELECT * FROM records WHERE {where} ORDER BY sheet_name, row_no", params
            ).fetchall()
        kind = imp["kind"] if imp else ""
        sheet_headers = None
        if kind == "comparison" and imp:
            metadata = json.loads(imp["metadata_json"])
            sheet_headers = {s["name"]: s["headers"] for s in metadata.get("sheets", [])}
        results = []
        for row in rows:
            item = dict(row)
            item["data"] = json.loads(item.pop("data_json"))
            if month is not None and self._record_month(kind, item["data"], sheet_headers) != month:
                continue
            results.append(item)
        if sort is not None:
            def sort_key(item: dict[str, Any]) -> tuple[int, Any]:
                value = item["data"].get(sort)
                number = _to_number(value)
                if number is not None:
                    return (0, number)
                return (1, str(value) if value is not None else "")
            results.sort(key=sort_key, reverse=(sort_dir == "desc"))
        return results[offset:offset + limit], len(results)

    def list_months(self, import_id: int) -> list[str]:
        return [entry["month"] for entry in self.month_summary(import_id)]

    def month_summary(self, import_id: int) -> list[dict[str, Any]]:
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT kind, metadata_json FROM imports WHERE id=?", (import_id,)).fetchone()
            if not imp:
                return []
            rows = db.execute("SELECT data_json FROM records WHERE import_id=?", (import_id,)).fetchall()
        kind = imp["kind"]
        sheet_headers = None
        if kind == "comparison":
            metadata = json.loads(imp["metadata_json"])
            sheet_headers = {s["name"]: s["headers"] for s in metadata.get("sheets", [])}
        counts: dict[str, int] = {}
        for row in rows:
            m = self._record_month(kind, json.loads(row["data_json"]), sheet_headers)
            if m:
                counts[m] = counts.get(m, 0) + 1
        return [{"month": m, "count": counts[m]} for m in sorted(counts)]

    def _comparison_sheet_headers_by_import(self) -> dict[int, dict[str, list]]:
        with closing(self.connect()) as db, db:
            rows = db.execute("SELECT id, metadata_json FROM imports WHERE kind='comparison'").fetchall()
        result = {}
        for row in rows:
            metadata = json.loads(row["metadata_json"])
            result[row["id"]] = {s["name"]: s["headers"] for s in metadata.get("sheets", [])}
        return result

    def comparison_month_summary(self) -> list[dict[str, Any]]:
        """相対表は古物台帳のように自動結合しないため、全ての取込を横断して月ごとに集計する。"""
        sheet_headers_by_import = self._comparison_sheet_headers_by_import()
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT import_id, data_json FROM records r JOIN imports i ON r.import_id=i.id WHERE i.kind='comparison'"
            ).fetchall()
        counts: dict[str, int] = {}
        for row in rows:
            data = json.loads(row["data_json"])
            headers = sheet_headers_by_import.get(row["import_id"])
            m = self._record_month("comparison", data, headers)
            if m:
                counts[m] = counts.get(m, 0) + 1
        return [{"month": m, "count": counts[m]} for m in sorted(counts)]

    def get_comparison_month_records(self, month: str, offset: int = 0,
                                     limit: int = 100) -> tuple[list[dict[str, Any]], int]:
        """指定した月（払出し側の年月日）の相対表データを、取込をまたいで横断的に取得する。"""
        sheet_headers_by_import = self._comparison_sheet_headers_by_import()
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT r.import_id, r.sheet_name, r.row_no, r.data_json, i.source_name "
                "FROM records r JOIN imports i ON r.import_id=i.id WHERE i.kind='comparison' "
                "ORDER BY r.import_id, r.sheet_name, r.row_no"
            ).fetchall()
        results = []
        for row in rows:
            data = json.loads(row["data_json"])
            headers = sheet_headers_by_import.get(row["import_id"])
            if self._record_month("comparison", data, headers) != month:
                continue
            item_headers = (headers or {}).get(row["sheet_name"])
            cells = [v for h, v in zip(item_headers, data["values"]) if h] if item_headers else data["values"]
            results.append({
                "import_id": row["import_id"], "source_name": row["source_name"],
                "sheet_name": row["sheet_name"], "row_no": row["row_no"], "cells": cells,
                "headers": [h for h in item_headers if h] if item_headers else None,
            })
        return results[offset:offset + limit], len(results)

    def find_comparison_duplicates(self) -> list[dict[str, Any]]:
        """相対表の複数の取込にまたがって、内容が完全に一致する行（同じファイルの重複取込など）
        をまとめる。同じ取込（同じファイル）内での重複は対象外（意図的に同じ内容の別取引である
        可能性が高く、件数も多くなりがちなため）。取込日時が古い方の行を削除候補として示す。
        """
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT r.import_id, r.sheet_name, r.row_no, r.data_json, i.source_name, i.imported_at "
                "FROM records r JOIN imports i ON r.import_id=i.id WHERE i.kind='comparison' "
                "ORDER BY r.import_id, r.sheet_name, r.row_no"
            ).fetchall()
        groups: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
        for row in rows:
            data = json.loads(row["data_json"])
            key = (row["sheet_name"], tuple(v for v in data["values"] if v is not None))
            groups.setdefault(key, []).append({
                "import_id": row["import_id"], "source_name": row["source_name"],
                "sheet": row["sheet_name"], "row_no": row["row_no"], "imported_at": row["imported_at"],
            })
        results = []
        for occurrences in groups.values():
            if len(occurrences) < 2:
                continue
            if len({o["import_id"] for o in occurrences}) < 2:
                continue
            # tie-break same-second imports by import_id, which is always increasing
            newest_key = max((o["imported_at"], o["import_id"]) for o in occurrences)
            for o in occurrences:
                o["recommended_delete"] = (o["imported_at"], o["import_id"]) != newest_key
            results.append({"occurrences": occurrences})
        return results

    def delete_comparison_record(self, import_id: int, sheet: str, row_no: int) -> None:
        with closing(self.connect()) as db, db:
            db.execute(
                "DELETE FROM allocations WHERE sale_import_id=? AND sale_sheet=? AND sale_row_no=?",
                (import_id, sheet, row_no),
            )
            db.execute(
                "DELETE FROM records WHERE import_id=? AND sheet_name=? AND row_no=?",
                (import_id, sheet, row_no),
            )

    def delete_recommended_comparison_duplicates(self) -> int:
        """find_comparison_duplicates() で「削除推奨（古い）」となった行をまとめて削除する。
        件数が多いときに1件ずつ削除する手間を省くための一括操作。
        """
        targets = [
            (o["import_id"], o["sheet"], o["row_no"])
            for group in self.find_comparison_duplicates()
            for o in group["occurrences"] if o["recommended_delete"]
        ]
        for import_id, sheet, row_no in targets:
            self.delete_comparison_record(import_id, sheet, row_no)
        return len(targets)

    def latest_merged_ledger_import(self) -> dict[str, Any] | None:
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT id, metadata_json FROM imports WHERE kind='ledger' ORDER BY id DESC"
            ).fetchall()
        for row in rows:
            if json.loads(row["metadata_json"]).get("source_format") == "merged":
                return self.get_import(row["id"])
        return None

    def list_exports(self) -> list[dict[str, Any]]:
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT id, import_id, template_id, mode, output_name, created_at FROM exports ORDER BY id DESC"
            ).fetchall()
        return [dict(row) for row in rows]

    def validate(self, import_id: int) -> list[CheckResult]:
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT * FROM imports WHERE id=?", (import_id,)).fetchone()
            if not imp:
                raise ValueError("取込IDが見つかりません")
            records = db.execute("SELECT * FROM records WHERE import_id=? ORDER BY sheet_name,row_no", (import_id,)).fetchall()
        if imp["kind"] == "ledger":
            return self._validate_ledger(records)
        if imp["kind"] == "inventory":
            return self._validate_inventory(records)
        if imp["kind"] == "export_data":
            return self._validate_export_data(records)
        checks = self._validate_comparison(records, json.loads(imp["metadata_json"]))
        checks.extend(self._validate_allocations(import_id))
        return checks

    def _validate_export_data(self, records: Iterable[sqlite3.Row]) -> list[CheckResult]:
        results: list[CheckResult] = []
        for record in records:
            data = json.loads(record["data_json"])
            row_no = record["row_no"]
            for key in ("年月日", "品名", "数量", "小計"):
                if data.get(key) in (None, ""):
                    results.append(CheckResult("REQUIRED", "error", f"必須項目「{key}」が空欄です", row_no=row_no))
            for key in ("金額", "数量", "小計"):
                if data.get(key) not in (None, "") and _to_number(data.get(key)) is None:
                    results.append(CheckResult("NUMBER_FORMAT", "error", f"{key}を数値として解釈できません", row_no=row_no))
        return results

    def _validate_inventory(self, records: Iterable[sqlite3.Row]) -> list[CheckResult]:
        results: list[CheckResult] = []
        for record in records:
            data = json.loads(record["data_json"])
            row_no = record["row_no"]
            for key in INVENTORY_COLUMNS:
                if data.get(key) in (None, ""):
                    results.append(CheckResult("REQUIRED", "error", f"必須項目「{key}」が空欄です", row_no=row_no))
            if data.get("仕入れ原価") not in (None, "") and _to_number(data.get("仕入れ原価")) is None:
                results.append(CheckResult("NUMBER_FORMAT", "error", "仕入れ原価を数値として解釈できません", row_no=row_no))
            if data.get("在庫数") not in (None, "") and _to_number(data.get("在庫数")) is None:
                results.append(CheckResult("NUMBER_FORMAT", "error", "在庫数を数値として解釈できません", row_no=row_no))
        return results

    def _validate_ledger(self, records: Iterable[sqlite3.Row]) -> list[CheckResult]:
        results: list[CheckResult] = []
        seen: set[tuple[Any, ...]] = set()
        for record in records:
            data = json.loads(record["data_json"])
            row_no = record["row_no"]
            for key in ("日時", "名前", "住所", "商品名", "個数", "金額"):
                if data.get(key) in (None, ""):
                    results.append(CheckResult("REQUIRED", "error", f"必須項目「{key}」が空欄です", row_no=row_no))
            try:
                qty, unit, amount = int(data.get("個数", 0)), float(str(data.get("単価", 0)).replace(",", "")), float(str(data.get("金額", 0)).replace(",", ""))
                if abs(qty * unit - amount) > 0.5:
                    results.append(CheckResult("AMOUNT_MISMATCH", "error", "数量×単価と金額が一致しません", row_no=row_no))
            except (TypeError, ValueError):
                results.append(CheckResult("NUMBER_FORMAT", "error", "数量・単価・金額を数値として解釈できません", row_no=row_no))
            key = tuple(data.get(k) for k in ("日時", "名前", "商品名", "個数", "金額"))
            if key in seen:
                results.append(CheckResult("DUPLICATE", "error", "同一取引の可能性があります", row_no=row_no))
            seen.add(key)
        return results

    def _validate_comparison(self, records: Iterable[sqlite3.Row], metadata: dict[str, Any]) -> list[CheckResult]:
        results: list[CheckResult] = []
        sheet_headers = {s["name"]: s["headers"] for s in metadata["sheets"]}
        for record in records:
            item = json.loads(record["data_json"]); values = item["values"]
            headers = sheet_headers[record["sheet_name"]]
            split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
            if split is None:
                results.append(CheckResult("LAYOUT", "error", "受入れ・払出しの境界を判定できません", record["sheet_name"], record["row_no"])); continue
            purchase, sale = values[:split], values[split:]
            if sale[0] not in (None, ""):
                purchase_qty_index = next((i for i, h in enumerate(headers[:split]) if h == "数量"), None)
                if purchase_qty_index is None or purchase[purchase_qty_index] in (None, ""):
                    results.append(CheckResult("PURCHASE_QTY", "error", "対応仕入数量が空欄です", record["sheet_name"], record["row_no"]))
                sale_qty_index = next((i for i, h in enumerate(headers[split:]) if h == "数量"), None)
                if purchase_qty_index is not None and sale_qty_index is not None:
                    pq, sq = purchase[purchase_qty_index], sale[sale_qty_index]
                    if not isinstance(pq, str) and not isinstance(sq, str) and pq != sq:
                        results.append(CheckResult("QTY_MISMATCH", "error", "販売数量と対応仕入数量が一致しません", record["sheet_name"], record["row_no"]))
        return results

    def _has_ledger_data(self) -> bool:
        with closing(self.connect()) as db, db:
            return db.execute("SELECT 1 FROM imports WHERE kind='ledger' LIMIT 1").fetchone() is not None

    def _validate_allocations(self, comparison_import_id: int) -> list[CheckResult]:
        if not self._has_ledger_data():
            return [CheckResult("LEDGER_MISSING", "error", "古物台帳が未取込のため、仕入との対応を確認できません")]
        checks: list[CheckResult] = []
        for result in self.allocate(comparison_import_id):
            if result["status"] == "matched" and result["note"]:
                checks.append(CheckResult("ALLOCATION_MISMATCH", "error", f"対応する仕入記録と{result['note']}",
                                          result["sheet"], result["row_no"]))
            elif result["status"] == "ambiguous":
                checks.append(CheckResult("ALLOCATION_AMBIGUOUS", "error",
                                          result["note"] or "対応する仕入の候補が複数あります。手動で選択してください",
                                          result["sheet"], result["row_no"]))
            elif result["status"] == "not_found":
                checks.append(CheckResult("ALLOCATION_NOT_FOUND", "error",
                                          result["note"] or "対応する仕入が古物台帳に見つかりません",
                                          result["sheet"], result["row_no"]))
        return checks

    def allocate(self, comparison_import_id: int) -> list[dict[str, Any]]:
        """既存の相対表データ（仕入・販売が同じ行に書かれている）の仕入側について、
        対応する古物台帳の購入記録を商品名・数量・日付・取引相手で自動的に探し、割り当てる。
        1件の古物台帳記録は、どこかひとつの販売にしか対応付けない（使い回さない）。
        """
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT * FROM imports WHERE id=?", (comparison_import_id,)).fetchone()
            if not imp or imp["kind"] != "comparison":
                raise ValueError("相対表の取込IDを指定してください")
            metadata = json.loads(imp["metadata_json"])
            records = db.execute(
                "SELECT * FROM records WHERE import_id=? ORDER BY sheet_name, row_no", (comparison_import_id,)
            ).fetchall()
            ledger_rows = db.execute(
                "SELECT r.id, r.data_json FROM records r JOIN imports i ON r.import_id = i.id WHERE i.kind='ledger'"
            ).fetchall()
            existing = {
                (row["sale_sheet"], row["sale_row_no"]): row
                for row in db.execute(
                    "SELECT * FROM allocations WHERE sale_import_id=?", (comparison_import_id,)
                ).fetchall()
            }
            other_consumed = {
                row["ledger_record_id"]
                for row in db.execute(
                    "SELECT ledger_record_id FROM allocations WHERE ledger_record_id IS NOT NULL "
                    "AND status IN ('matched','manual') AND sale_import_id != ?",
                    (comparison_import_id,),
                ).fetchall()
            }

        sheet_headers = {s["name"]: s["headers"] for s in metadata["sheets"]}
        ledger_by_id: dict[int, dict[str, Any]] = {}
        for row in ledger_rows:
            data = json.loads(row["data_json"])
            ledger_by_id[row["id"]] = {
                "id": row["id"],
                "name": (data.get("名前") or "").strip(),
                "product": (data.get("商品名") or "").strip(),
                "qty": _to_number(data.get("個数")),
                "date": _to_date(data.get("日時")),
                "amount": _to_number(data.get("金額")),
            }

        consumed = set(other_consumed)
        for row in existing.values():
            if row["status"] == "manual" and row["ledger_record_id"] is not None:
                consumed.add(row["ledger_record_id"])

        results: list[dict[str, Any]] = []
        upserts: list[tuple[Any, ...]] = []
        now = datetime.now().isoformat(timespec="seconds")

        for record in records:
            headers = sheet_headers.get(record["sheet_name"], [])
            split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
            if split is None:
                continue
            values = json.loads(record["data_json"])["values"]
            purchase, sale = values[:split], values[split:]
            if sale[0] in (None, ""):
                continue
            purchase_headers = headers[:split]

            prior = existing.get((record["sheet_name"], record["row_no"]))
            if prior and prior["status"] == "manual":
                results.append({
                    "row_no": record["row_no"], "sheet": record["sheet_name"], "status": "manual",
                    "ledger_record_id": prior["ledger_record_id"],
                    "ledger": ledger_by_id.get(prior["ledger_record_id"]),
                    "product": None, "note": "手動で割当済み", "candidates": [],
                })
                continue

            product_idx = _index_of(purchase_headers, "品目")
            qty_idx = _index_of(purchase_headers, "数量")
            name_idx = _index_of(purchase_headers, "相手方名")
            amount_idx = _index_of(purchase_headers, "代価", contains=True)
            product = str(purchase[product_idx]).strip() if product_idx is not None and purchase[product_idx] not in (None, "") else None
            qty = _to_number(purchase[qty_idx]) if qty_idx is not None else None
            purchase_date = _to_date(purchase[0])
            counterparty = str(purchase[name_idx]).strip() if name_idx is not None and purchase[name_idx] not in (None, "") else None
            amount = _to_number(purchase[amount_idx]) if amount_idx is not None else None

            candidates = [c for c in ledger_by_id.values() if c["id"] not in consumed and product and c["product"] == product]
            strict = [c for c in candidates if c["qty"] == qty and c["date"] == purchase_date]
            pool = strict or candidates
            if counterparty:
                named = [c for c in pool if c["name"] == counterparty]
                if named:
                    pool = named

            status: str; ledger_id: int | None = None; note: str | None; candidate_ids: list[int] = []
            if not product:
                status, note = "not_found", "品目が空欄です"
            elif len(pool) == 1:
                status, ledger_id = "matched", pool[0]["id"]
                consumed.add(ledger_id)
                mismatches = []
                if qty is not None and pool[0]["qty"] != qty:
                    mismatches.append("数量")
                if purchase_date is not None and pool[0]["date"] != purchase_date:
                    mismatches.append("日付")
                if amount is not None and pool[0]["amount"] is not None and abs(pool[0]["amount"] - amount) > 0.5:
                    mismatches.append("金額")
                note = ("・".join(mismatches) + "が一致しません") if mismatches else None
            elif len(pool) > 1:
                status, note, candidate_ids = "ambiguous", f"候補{len(pool)}件から選択してください", [c["id"] for c in pool]
            else:
                status, note, candidate_ids = "not_found", "一致する古物台帳の記録がありません", [c["id"] for c in candidates]

            upserts.append((comparison_import_id, record["sheet_name"], record["row_no"], ledger_id, status,
                            json_text(candidate_ids), note, now))
            results.append({
                "row_no": record["row_no"], "sheet": record["sheet_name"], "status": status,
                "ledger_record_id": ledger_id, "ledger": ledger_by_id.get(ledger_id), "product": product,
                "note": note, "candidates": [ledger_by_id[cid] for cid in candidate_ids],
            })

        with closing(self.connect()) as db, db:
            db.executemany(
                """INSERT INTO allocations(sale_import_id, sale_sheet, sale_row_no, ledger_record_id, status, candidates_json, note, created_at)
                   VALUES(?,?,?,?,?,?,?,?)
                   ON CONFLICT(sale_import_id, sale_sheet, sale_row_no) DO UPDATE SET
                     ledger_record_id=excluded.ledger_record_id, status=excluded.status,
                     candidates_json=excluded.candidates_json, note=excluded.note, created_at=excluded.created_at
                """,
                upserts,
            )
        return results

    def search_ledger(self, query: str, limit: int = 30) -> list[dict[str, Any]]:
        query = query.strip()
        if not query:
            return []
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT r.id, r.data_json FROM records r JOIN imports i ON r.import_id=i.id WHERE i.kind='ledger' ORDER BY r.id"
            ).fetchall()
            consumed = {
                row["ledger_record_id"]
                for row in db.execute(
                    "SELECT ledger_record_id FROM allocations WHERE ledger_record_id IS NOT NULL AND status IN ('matched','manual')"
                ).fetchall()
            }
        result = []
        for row in rows:
            if row["id"] in consumed:
                continue
            data = json.loads(row["data_json"])
            haystack = f"{data.get('商品名', '')} {data.get('名前', '')}"
            if query not in haystack:
                continue
            result.append({"id": row["id"], "name": data.get("名前"), "product": data.get("商品名"),
                           "qty": data.get("個数"), "date": data.get("日時"), "amount": data.get("金額")})
            if len(result) >= limit:
                break
        return result

    def set_manual_allocation(self, comparison_import_id: int, sheet: str, row_no: int,
                              ledger_record_id: int | None) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self.connect()) as db, db:
            db.execute(
                """INSERT INTO allocations(sale_import_id, sale_sheet, sale_row_no, ledger_record_id, status, candidates_json, note, created_at)
                   VALUES(?,?,?,?,?,?,?,?)
                   ON CONFLICT(sale_import_id, sale_sheet, sale_row_no) DO UPDATE SET
                     ledger_record_id=excluded.ledger_record_id, status='manual', candidates_json='[]',
                     note=NULL, created_at=excluded.created_at
                """,
                (comparison_import_id, sheet, row_no, ledger_record_id, "manual", "[]", None, now),
            )

    def _inventory_for_month(self, month: str | None) -> dict[str, float]:
        """指定した基準年月（例: "2026-04"）と一致する期末在庫表から、商品名と仕入れ原価の
        対応を返す。より新しい基準月の在庫表を代わりに使うと、対象月にはまだ仕入れていない
        商品が紛れ込むおそれがあるため、必ず同じ基準月の在庫表のみを使う。該当する在庫表が
        なければ空の辞書を返す（呼び出し側は「その月の在庫表がない」ものとして扱う）。
        """
        if month is None:
            return {}
        with closing(self.connect()) as db, db:
            imps = db.execute(
                "SELECT id, imported_at, metadata_json FROM imports WHERE kind='inventory'"
            ).fetchall()
            matches = [imp for imp in imps if json.loads(imp["metadata_json"]).get("as_of") == month]
            if not matches:
                return {}
            imp = max(matches, key=lambda row: row["imported_at"])
            rows = db.execute("SELECT data_json FROM records WHERE import_id=?", (imp["id"],)).fetchall()
        costs: dict[str, float] = {}
        for row in rows:
            data = json.loads(row["data_json"])
            product = (data.get("商品名") or "").strip()
            cost = _to_number(data.get("仕入れ原価"))
            if product and cost is not None:
                costs[product] = cost
        return costs

    def search_inventory(self, query: str, month: str | None, limit: int = 30) -> list[dict[str, Any]]:
        query = query.strip()
        if not query:
            return []
        result = []
        for product, cost in self._inventory_for_month(month).items():
            if query in product:
                result.append({"product": product, "unit_cost": cost})
                if len(result) >= limit:
                    break
        return result

    def propose_ledger_breakdown(self, ledger_import_id: int, month: str | None = None) -> list[dict[str, Any]]:
        """商品名が空欄の古物台帳の行について、相対表から判明している内訳（確定）と、
        まだ手動で埋めていない残額（要確認）を計算する。金額の内訳は書き換えず、
        常にこの場で再計算するため、相対表・在庫表を取込み直すと結果も更新される。
        month（例: "2026-04"）を指定すると、その仕入年月（日時）の行だけに絞り込む。
        期末在庫表は各行の仕入年月と同じ基準月のものだけを使う。別の月の在庫表を代わりに
        使うと、その月にはまだ仕入れていない商品が紛れ込むおそれがあるため。
        """
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT * FROM imports WHERE id=?", (ledger_import_id,)).fetchone()
            if not imp or imp["kind"] != "ledger":
                raise ValueError("古物台帳の取込IDを指定してください")
            ledger_records = db.execute(
                "SELECT * FROM records WHERE import_id=? ORDER BY row_no", (ledger_import_id,)
            ).fetchall()
            comparison_rows = db.execute(
                "SELECT r.import_id, r.sheet_name, r.data_json FROM records r "
                "JOIN imports i ON r.import_id = i.id WHERE i.kind='comparison'"
            ).fetchall()
            comparison_metadata = {
                row["id"]: json.loads(row["metadata_json"])
                for row in db.execute("SELECT id, metadata_json FROM imports WHERE kind='comparison'").fetchall()
            }
            manual_items = db.execute(
                "SELECT * FROM ledger_items WHERE ledger_import_id=? ORDER BY id", (ledger_import_id,)
            ).fetchall()

        inventory_cache: dict[str | None, dict[str, float]] = {}

        def inventory_for(row_month: str | None) -> dict[str, float]:
            if row_month not in inventory_cache:
                inventory_cache[row_month] = self._inventory_for_month(row_month)
            return inventory_cache[row_month]

        comparison_by_key: dict[tuple[str, date | None], list[dict[str, Any]]] = {}
        for row in comparison_rows:
            metadata = comparison_metadata.get(row["import_id"], {})
            headers = next((s["headers"] for s in metadata.get("sheets", []) if s["name"] == row["sheet_name"]), None)
            if not headers:
                continue
            split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
            if split is None:
                continue
            values = json.loads(row["data_json"])["values"]
            purchase, sale = values[:split], values[split:]
            if sale[0] in (None, ""):
                continue
            purchase_headers = headers[:split]
            product_idx = _index_of(purchase_headers, "品目")
            qty_idx = _index_of(purchase_headers, "数量")
            name_idx = _index_of(purchase_headers, "相手方名")
            if product_idx is None or name_idx is None:
                continue
            product = str(purchase[product_idx]).strip() if purchase[product_idx] not in (None, "") else None
            counterparty = str(purchase[name_idx]).strip() if purchase[name_idx] not in (None, "") else None
            qty = _to_number(purchase[qty_idx]) if qty_idx is not None else None
            purchase_date = _to_date(purchase[0])
            if not product or not counterparty:
                continue
            comparison_by_key.setdefault((counterparty, purchase_date), []).append({"product": product, "qty": qty or 0})

        manual_by_row: dict[int, list[dict[str, Any]]] = {}
        for item in manual_items:
            manual_by_row.setdefault(item["ledger_row_no"], []).append(dict(item))

        results: list[dict[str, Any]] = []
        for record in ledger_records:
            data = json.loads(record["data_json"])
            if data.get("商品名") not in (None, ""):
                continue
            row_no = record["row_no"]
            name = (data.get("名前") or "").strip()
            purchase_date = _to_date(data.get("日時"))
            row_month = f"{purchase_date.year:04d}-{purchase_date.month:02d}" if purchase_date else None
            if month is not None and row_month != month:
                continue
            total = _to_number(data.get("金額"))
            inventory = inventory_for(row_month)

            known = []
            for entry in comparison_by_key.get((name, purchase_date), []):
                cost = inventory.get(entry["product"])
                qty = entry["qty"]
                known.append({
                    "product": entry["product"], "qty": qty, "unit_cost": cost,
                    "amount": (cost * qty) if cost is not None else None, "source": "comparison",
                })
            manual = [{
                "id": item["id"], "product": item["product"], "qty": item["qty"],
                "unit_cost": item["unit_cost"], "amount": item["amount"], "source": "manual",
            } for item in manual_by_row.get(row_no, [])]

            accounted = sum((i["amount"] or 0) for i in known + manual)
            remainder = (total - accounted) if total is not None else None
            results.append({
                "row_no": row_no, "name": name, "date": purchase_date, "month": row_month, "total": total,
                "known_items": known, "manual_items": manual, "remainder": remainder,
                "resolved": remainder is not None and abs(remainder) < 0.5,
                "inventory_available": bool(inventory),
            })
        return results

    def suggest_ledger_completion(self, ledger_import_id: int, row_no: int) -> list[dict[str, Any]] | None:
        entry = next(
            (b for b in self.propose_ledger_breakdown(ledger_import_id) if b["row_no"] == row_no), None,
        )
        if not entry or entry["resolved"] or entry["remainder"] is None:
            return None
        return suggest_ledger_items(entry["remainder"], self._inventory_for_month(entry["month"]))

    def add_ledger_item(self, ledger_import_id: int, row_no: int, product: str, qty: float) -> None:
        with closing(self.connect()) as db, db:
            record = db.execute(
                "SELECT data_json FROM records WHERE import_id=? AND row_no=?", (ledger_import_id, row_no)
            ).fetchone()
        if not record:
            raise ValueError("対象の行が見つかりません")
        purchase_date = _to_date(json.loads(record["data_json"]).get("日時"))
        month = f"{purchase_date.year:04d}-{purchase_date.month:02d}" if purchase_date else None
        inventory = self._inventory_for_month(month)
        if not inventory:
            raise ValueError(f"{month or '該当行の仕入年月'}の期末在庫表が見つかりません。先に同じ基準月の期末在庫表を取り込んでください")
        cost = inventory.get(product)
        if cost is None:
            raise ValueError(f"{month}の期末在庫表に「{product}」の仕入れ原価が見つかりません")
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self.connect()) as db, db:
            db.execute(
                "INSERT INTO ledger_items(ledger_import_id, ledger_row_no, product, qty, unit_cost, amount, source, created_at) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (ledger_import_id, row_no, product, qty, cost, cost * qty, "manual", now),
            )

    def remove_ledger_item(self, item_id: int) -> None:
        with closing(self.connect()) as db, db:
            db.execute("DELETE FROM ledger_items WHERE id=? AND source='manual'", (item_id,))

    @staticmethod
    def _ledger_duplicate_key(data: dict[str, Any]) -> str:
        return json.dumps([data.get(k) for k in ("日時", "名前", "商品名", "個数", "金額")], ensure_ascii=False)

    def find_ledger_duplicates(self, ledger_import_id: int) -> list[dict[str, Any]]:
        """指定した古物台帳の取込（通常は結合済みのもの）の中で、日時・名前・商品名・
        個数・金額が完全に一致する行をまとめる。結合の際にも重複候補として報告されるが、
        取込直後に見落とした場合や、後から見直したい場合のためにいつでも確認できるようにする。
        本人確認情報（ふりがな・生年月日・住所・電話番号）がより多く埋まっている方を残す候補とし、
        情報が少ない方に削除推奨を付ける。情報量が同じでも行の中身が完全に一致するなら
        （＝どちらを残しても情報が減らない）1件だけ残して削除推奨を付ける。情報量が同じで
        中身にも違いがある場合のみ、どちらも推奨しない＝人が判断する。
        いずれの場合も「残す（重複ではない）」を選べば、削除推奨に関わらず次回以降ここに
        出てこないようにできる（同額・別商品の取引を偶然2回行った、など）。
        """
        with closing(self.connect()) as db, db:
            rows = db.execute(
                "SELECT row_no, data_json FROM records WHERE import_id=? ORDER BY row_no", (ledger_import_id,)
            ).fetchall()
            dismissed = {
                r["key_json"] for r in db.execute(
                    "SELECT key_json FROM ledger_duplicate_dismissals WHERE import_id=?", (ledger_import_id,)
                ).fetchall()
            }
        groups: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            data = json.loads(row["data_json"])
            key = self._ledger_duplicate_key(data)
            completeness = sum(1 for f in ("ふりがな", "生年月日", "住所", "電話番号") if data.get(f))
            groups.setdefault(key, []).append({"row_no": row["row_no"], "data": data, "completeness": completeness})
        results = []
        for key, occurrences in groups.items():
            if len(occurrences) < 2 or key in dismissed:
                continue
            best = max(o["completeness"] for o in occurrences)
            clear_winner = sum(1 for o in occurrences if o["completeness"] == best) == 1
            if clear_winner:
                for o in occurrences:
                    o["recommended_delete"] = o["completeness"] < best
            elif len({json.dumps(o["data"], sort_keys=True) for o in occurrences}) == 1:
                for index, o in enumerate(occurrences):
                    o["recommended_delete"] = index > 0
            else:
                for o in occurrences:
                    o["recommended_delete"] = False
            results.append({"occurrences": occurrences})
        return results

    def dismiss_ledger_duplicate(self, import_id: int, row_no: int) -> None:
        """この行が属する重複候補グループを「重複ではない（両方残す）」として、
        以後 find_ledger_duplicates の一覧に出てこないようにする。
        """
        with closing(self.connect()) as db, db:
            row = db.execute(
                "SELECT data_json FROM records WHERE import_id=? AND row_no=?", (import_id, row_no)
            ).fetchone()
            if not row:
                raise ValueError("対象の行が見つかりません")
            key = self._ledger_duplicate_key(json.loads(row["data_json"]))
            now = datetime.now().isoformat(timespec="seconds")
            db.execute(
                "INSERT OR IGNORE INTO ledger_duplicate_dismissals(import_id, key_json, created_at) VALUES(?,?,?)",
                (import_id, key, now),
            )

    def delete_ledger_record(self, import_id: int, row_no: int) -> None:
        with closing(self.connect()) as db, db:
            db.execute("DELETE FROM ledger_items WHERE ledger_import_id=? AND ledger_row_no=?", (import_id, row_no))
            db.execute("DELETE FROM records WHERE import_id=? AND row_no=?", (import_id, row_no))

    def delete_recommended_ledger_duplicates(self, ledger_import_id: int) -> int:
        """find_ledger_duplicates() で削除推奨となった行をまとめて削除する。"""
        targets = [
            o["row_no"]
            for group in self.find_ledger_duplicates(ledger_import_id)
            for o in group["occurrences"] if o["recommended_delete"]
        ]
        for row_no in targets:
            self.delete_ledger_record(ledger_import_id, row_no)
        return len(targets)

    def delete_import(self, import_id: int) -> None:
        """取込を削除する（誤って同じファイルを2回取り込んだ場合などの手動クリーンアップ用）。
        紐づく明細・手動内訳・仕入対応も合わせて削除する。過去の出力履歴（exports）は
        監査証跡として残す。
        """
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT id FROM imports WHERE id=?", (import_id,)).fetchone()
            if not imp:
                raise ValueError("取込IDが見つかりません")
            db.execute("DELETE FROM ledger_items WHERE ledger_import_id=?", (import_id,))
            db.execute("DELETE FROM allocations WHERE sale_import_id=?", (import_id,))
            db.execute("DELETE FROM records WHERE import_id=?", (import_id,))
            db.execute("DELETE FROM imports WHERE id=?", (import_id,))

    def delete_imports(self, import_ids: list[int]) -> int:
        """複数の取込をまとめて削除する（一覧画面でのチェックボックス選択削除用）。
        既に無い取込IDは無視して先へ進む。実際に削除できた件数を返す。
        """
        count = 0
        for import_id in import_ids:
            try:
                self.delete_import(import_id)
                count += 1
            except ValueError:
                continue
        return count

    def export_completed_ledger(self, ledger_import_id: int, output: str | Path) -> Path:
        breakdown = self.propose_ledger_breakdown(ledger_import_id)
        unresolved = [b for b in breakdown if not b["resolved"]]
        if unresolved:
            raise ValueError(f"{len(unresolved)}件の行で内訳の金額が合計と一致していません。先に確定してください")
        breakdown_by_row = {b["row_no"]: b for b in breakdown}
        with closing(self.connect()) as db, db:
            records = db.execute(
                "SELECT * FROM records WHERE import_id=? ORDER BY row_no", (ledger_import_id,)
            ).fetchall()
        output = Path(output).resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        with output.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            for record in records:
                data = json.loads(record["data_json"])
                entry = breakdown_by_row.get(record["row_no"])
                if not entry:
                    writer.writerow(data)
                    continue
                original_total = data.get("金額")
                for item in entry["known_items"] + entry["manual_items"]:
                    row = dict(data)
                    row["商品名"] = item["product"]
                    row["個数"] = item["qty"]
                    row["単価"] = item["unit_cost"]
                    row["金額"] = item["amount"]
                    note = f"内訳復元（元合計{original_total}円・{'相対表' if item['source'] == 'comparison' else '手動'}）"
                    row["備考"] = f"{data.get('備考') or ''} {note}".strip()
                    writer.writerow(row)
        return output

    def merge_ledger_exports(self, import_ids: list[int]) -> dict[str, Any]:
        """複数の古物台帳の取込（形式の異なる取込元でもよい）を1つにまとめる。
        日時・名前・商品名・個数・金額が完全に一致する行は重複の可能性として報告する
        （自動では除外しない。実際に重複かどうかは人が確認する）。
        """
        if not import_ids:
            raise ValueError("結合する取込を1件以上選択してください")
        with closing(self.connect()) as db, db:
            placeholders = ",".join("?" * len(import_ids))
            imps = db.execute(f"SELECT * FROM imports WHERE id IN ({placeholders})", import_ids).fetchall()
            if len(imps) != len(set(import_ids)) or any(i["kind"] != "ledger" for i in imps):
                raise ValueError("古物台帳の取込のみ選択できます")
            records = db.execute(
                f"SELECT * FROM records WHERE import_id IN ({placeholders}) ORDER BY import_id, row_no",
                import_ids,
            ).fetchall()
        rows: list[dict[str, Any]] = []
        seen: dict[tuple[Any, ...], int] = {}
        duplicates: list[dict[str, Any]] = []
        for record in records:
            data = json.loads(record["data_json"])
            rows.append(data)
            key = tuple(data.get(k) for k in ("日時", "名前", "商品名", "個数", "金額"))
            if key in seen:
                duplicates.append({"日時": data.get("日時"), "名前": data.get("名前"),
                                   "商品名": data.get("商品名"), "import_ids": [seen[key], record["import_id"]]})
            seen[key] = record["import_id"]
        rows.sort(key=lambda r: r.get("日時") or "")
        return {"rows": rows, "duplicates": duplicates}

    def merge_ledger_imports(self, import_ids: list[int]) -> dict[str, Any]:
        """複数の古物台帳の取込を、新しい1件の古物台帳の取込としてまとめる。
        内訳復元・チェックをまとめて1回で行えるよう、取込直後（エラーが残っていて）でも結合できる。
        元の取込は削除せず、参照用にそのまま残す。
        """
        with closing(self.connect()) as db, db:
            placeholders = ",".join("?" * len(import_ids))
            imps = db.execute(f"SELECT * FROM imports WHERE id IN ({placeholders})", import_ids).fetchall()
        merged = self.merge_ledger_exports(import_ids)
        names = "・".join(imp["source_name"] for imp in imps)
        import_id = self._save_import(
            "ledger", f"結合（{names}）", "merged", merged["rows"],
            {"source_format": "merged", "merged_from": import_ids},
        )
        return {"import_id": import_id, "total": len(merged["rows"]), "duplicates": merged["duplicates"]}

    def auto_merge_ledger_imports(self) -> dict[str, Any]:
        """未結合の古物台帳の取込（まだどの結合にも含まれていない生データ）をすべて自動で結合する。
        どれとどれを結合するかを選ぶ必要はない。既存の結合結果は対象に含めない（重複結合を防ぐため）。
        """
        with closing(self.connect()) as db, db:
            rows = db.execute("SELECT id, metadata_json FROM imports WHERE kind='ledger'").fetchall()
        raw_ids = [
            row["id"] for row in rows
            if json.loads(row["metadata_json"]).get("source_format") != "merged"
        ]
        if len(raw_ids) < 2:
            raise ValueError("結合できる古物台帳の取込が2件未満です")
        return self.merge_ledger_imports(raw_ids)

    def build_comparison(self, export_data_import_id: int, template_id: int) -> dict[str, Any]:
        """輸出データ（払出し側のみ）と古物台帳から、指定テンプレートの構造に沿った
        相対表（輸出販売シート）を組み立てる。対応する仕入は品名・数量が一致し、
        輸出データの年月日以前の古物台帳記録から、先入先出（最も古い未使用のもの）で
        自動的に対応付ける。対応が見つからない行は品目だけ埋めて残りは空欄にし、
        既存のチェック・仕入対応の確認画面でエラーとして検出・手動対応できるようにする。
        """
        with closing(self.connect()) as db, db:
            exp_imp = db.execute("SELECT * FROM imports WHERE id=?", (export_data_import_id,)).fetchone()
            if not exp_imp or exp_imp["kind"] != "export_data":
                raise ValueError("輸出データの取込IDを指定してください")
            template = db.execute("SELECT * FROM template_versions WHERE id=?", (template_id,)).fetchone()
            if not template or template["report_type"] != "comparison":
                raise ValueError("相対表のテンプレートを指定してください")
            export_records = db.execute(
                "SELECT * FROM records WHERE import_id=? ORDER BY row_no", (export_data_import_id,)
            ).fetchall()
            ledger_rows = db.execute(
                "SELECT r.id, r.data_json FROM records r JOIN imports i ON r.import_id = i.id WHERE i.kind='ledger'"
            ).fetchall()
            other_consumed = {
                row["ledger_record_id"]
                for row in db.execute(
                    "SELECT ledger_record_id FROM allocations WHERE ledger_record_id IS NOT NULL AND status IN ('matched','manual')"
                ).fetchall()
            }

        settings = json.loads(template["settings_json"])
        sheet_meta = next((s for s in settings.get("sheets", []) if s["name"] == "輸出販売"), None)
        if sheet_meta is None:
            raise ValueError("このテンプレートに「輸出販売」シートが見つかりません")
        headers = sheet_meta["headers"]
        split = next((i for i, h in enumerate(headers) if i > 0 and h == "年月日"), None)
        if split is None:
            raise ValueError("テンプレートの受入れ・払出しの境界を判定できません")
        purchase_headers, sale_headers = headers[:split], headers[split:]

        p_product = _index_of(purchase_headers, "品目")
        p_qty = _index_of(purchase_headers, "数量")
        p_amount = _index_of(purchase_headers, "代価", contains=True)
        p_unit = _index_of(purchase_headers, "単価", contains=True)
        p_name = _index_of(purchase_headers, "相手方名")

        s_qty = _index_of(sale_headers, "数量")
        s_amount = _index_of(sale_headers, "代価", contains=True)
        s_unit = _index_of(sale_headers, "単価", contains=True)
        s_name = _index_of(sale_headers, "相手方名")
        s_payment = _index_of(sale_headers, "支払方法")
        s_currency = _index_of(sale_headers, "通貨")

        ledger_pool = []
        for row in ledger_rows:
            if row["id"] in other_consumed:
                continue
            data = json.loads(row["data_json"])
            ledger_pool.append({
                "id": row["id"], "name": (data.get("名前") or "").strip(),
                "product": (data.get("商品名") or "").strip(),
                "qty": _to_number(data.get("個数")), "date": _to_date(data.get("日時")),
                "unit": _to_number(data.get("単価")), "amount": _to_number(data.get("金額")),
            })

        consumed: set[int] = set()
        rows_out: list[dict[str, Any]] = []
        allocation_rows: list[tuple[Any, ...]] = []
        unmatched = 0
        now = datetime.now().isoformat(timespec="seconds")

        for record in export_records:
            data = json.loads(record["data_json"])
            product = (data.get("品名") or "").strip()
            qty = _to_number(data.get("数量"))
            sale_date = _to_date(data.get("年月日"))

            candidates = [
                c for c in ledger_pool
                if c["id"] not in consumed and product and c["product"] == product and c["qty"] == qty
                and (sale_date is None or c["date"] is None or c["date"] <= sale_date)
            ]
            candidates.sort(key=lambda c: c["date"] or date.max)
            match = candidates[0] if candidates else None
            if match:
                consumed.add(match["id"])
            else:
                unmatched += 1

            values: list[Any] = [None] * len(headers)
            if p_product is not None:
                values[p_product] = product
            if match:
                values[0] = match["date"]
                if p_qty is not None: values[p_qty] = match["qty"]
                if p_unit is not None: values[p_unit] = match["unit"]
                if p_amount is not None: values[p_amount] = match["amount"]
                if p_name is not None: values[p_name] = match["name"]
            values[split] = sale_date
            if s_qty is not None: values[split + s_qty] = qty
            if s_unit is not None: values[split + s_unit] = _to_number(data.get("金額"))
            if s_amount is not None: values[split + s_amount] = _to_number(data.get("小計"))
            if s_name is not None: values[split + s_name] = data.get("相手方名")
            if s_payment is not None: values[split + s_payment] = data.get("支払方法")
            if s_currency is not None: values[split + s_currency] = data.get("通貨")

            rows_out.append({"sheet": "輸出販売", "row_no": record["row_no"], "values": values})
            if match:
                allocation_rows.append(("輸出販売", record["row_no"], match["id"], "matched", "[]", None, now))

        metadata = {
            "sheets": [{"name": "輸出販売", "headers": headers, "max_column": len(headers)}],
            "built_from": {"export_data_import_id": export_data_import_id, "template_id": template_id},
        }
        import_id = self._save_import(
            "comparison", f"{exp_imp['source_name']}（組立）", exp_imp["source_sha256"], rows_out, metadata,
        )
        if allocation_rows:
            with closing(self.connect()) as db, db:
                db.executemany(
                    """INSERT INTO allocations(sale_import_id, sale_sheet, sale_row_no, ledger_record_id, status, candidates_json, note, created_at)
                       VALUES(?,?,?,?,?,?,?,?)
                       ON CONFLICT(sale_import_id, sale_sheet, sale_row_no) DO UPDATE SET
                         ledger_record_id=excluded.ledger_record_id, status=excluded.status,
                         candidates_json=excluded.candidates_json, note=excluded.note, created_at=excluded.created_at
                    """,
                    [(import_id, *row) for row in allocation_rows],
                )
        return {"import_id": import_id, "total": len(rows_out), "unmatched": unmatched}

    def export(self, import_id: int, output: str | Path, template_id: int | None = None,
               preview: bool = False, month: str | None = None) -> Path:
        with closing(self.connect()) as db, db:
            imp = db.execute("SELECT * FROM imports WHERE id=?", (import_id,)).fetchone()
            if not imp:
                raise ValueError("取込IDが見つかりません")
            records = db.execute("SELECT * FROM records WHERE import_id=? ORDER BY sheet_name,row_no", (import_id,)).fetchall()
            template = db.execute("SELECT * FROM template_versions WHERE id=?", (template_id,)).fetchone() if template_id else None

        if month is not None:
            if imp["kind"] != "ledger":
                raise ValueError("月ごとの出力は古物台帳のみ対応しています")
            records = [r for r in records if self._record_month("ledger", json.loads(r["data_json"]), None) == month]
            if not records:
                raise ValueError(f"{month} の古物台帳データがありません")
            checks = self._validate_ledger(records)
        else:
            checks = self._validate_ledger(records) if imp["kind"] == "ledger" else self.validate(import_id)

        if any(c.level == "error" for c in checks) and not preview:
            raise ValueError("検証エラーがあるため正式出力できません。--preview を指定すると確認用出力が可能です")
        output = Path(output).resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        if imp["kind"] == "ledger":
            self._export_ledger(records, output)
        else:
            if not template: raise ValueError("相対表出力には --template-id が必要です")
            self._export_comparison(records, Path(template["stored_path"]), output, preview)
        with closing(self.connect()) as db, db:
            db.execute("INSERT INTO exports(import_id,template_id,mode,output_name,output_sha256,checks_json,created_at) VALUES(?,?,?,?,?,?,?)",
                       (import_id, template_id, "preview" if preview else "formal", output.name, sha256(output),
                        json_text([c.__dict__ for c in checks]), datetime.now().isoformat(timespec="seconds")))
        return output

    def _export_ledger(self, records: Iterable[sqlite3.Row], output: Path) -> None:
        with output.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=LEDGER_COLUMNS, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            for record in records: writer.writerow(json.loads(record["data_json"]))

    def _export_comparison(self, records: Iterable[sqlite3.Row], template: Path, output: Path, preview: bool) -> None:
        wb = load_workbook(template, read_only=False, data_only=False)
        grouped: dict[str, list[dict[str, Any]]] = {}
        for record in records: grouped.setdefault(record["sheet_name"], []).append(json.loads(record["data_json"]))
        for ws in wb.worksheets:
            items = grouped.get(ws.title, [])
            for r in range(3, ws.max_row + 1):
                for c in range(1, ws.max_column + 1): ws.cell(r, c).value = None
            for index, item in enumerate(items, 3):
                for col, value in enumerate(item["values"], 1): ws.cell(index, col).value = value
            if preview:
                ws.oddHeader.center.text = "確認用（正式帳票ではありません）"
        fd, temp_name = tempfile.mkstemp(suffix=".xlsx", dir=output.parent); os.close(fd)
        try:
            wb.save(temp_name); os.replace(temp_name, output)
        finally:
            if os.path.exists(temp_name): os.unlink(temp_name)
